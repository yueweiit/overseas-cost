"""中文用途：解析成本总表/钉钉附件 xlsx 为现有导入 block 结构。"""

from __future__ import annotations

import re
from datetime import date, datetime, time
from pathlib import Path
from typing import Any

from overseas_costing.utils.field_mapper import normalize_unit

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - 只在真实解析 xlsx 时才需要报错
    load_workbook = None


MAX_EXCEL_COLUMN = "BE"

ITEM_COLUMNS = ["A", "B", "C", "D", "E", "F", "G", "H", "BA", "BC", "T"]

EXTRA_COLUMN_MAP = {
    "N": "ccRate",
    "O": "ccAntiDumping",
    "P": "igiRate",
    "Q": "igiAmount",
    "R": "ivaRate",
    "S": "ivaAmount",
    "U": "dta",
    "V": "prvDuty",
    "W": "prvIva",
    "X": "importTaxTotal",
    "Y": "revalidacion",
    "Z": "maniobras",
    "AA": "muellaje",
    "AB": "entregaMercancia",
    "AC": "previo",
    "AD": "serviceAA",
    "AE": "almacenajes",
    "AF": "reconocimientoAduanero",
    "AG": "honorarios",
    "AH": "complementoManiobras",
    "AI": "desconsolidacion",
    "AJ": "maniobraFalso",
    "AK": "arrastre",
    "AL": "patioRegulador",
    "AM": "entregaVacio",
    "AN": "limpiezaContenedor",
    "AO": "mexicoCustomsMxn",
    "AP": "mexicoCustomsRmb",
    "AQ": "mexicoCustomsUsd",
    "AR": "mexicoInlandMxn",
    "AS": "mexicoMiscMxn",
    "AT": "mexicoInlandMiscRmb",
    "AU": "chinaToMexicoFreightRmb",
    "AV": "grossWeightKg",
    "AW": "weightRatio",
    "AX": "freightAllocRmb",
    "AY": "freightAllocMxn",
    "AZ": "totalLogisticsMxn",
    "BB": "totalCostRmb",
    "BD": "projectCollection",
    "BE": "transportMode",
}

BLOCK_COLUMN_MAP = {
    "customsNo": "I",
    "waybillNo": "J",
    "chinaMiscRmb": "K",
    "chinaMiscMxn": "L",
    "oceanUsd": "M",
    "mexicoInlandMxn": "AR",
    "mexicoMiscMxn": "AS",
    "mexicoInlandMiscRmb": "AT",
    "chinaToMexicoFreightRmb": "AU",
    "projectCollection": "BD",
    "transportMode": "BE",
}


def col_to_index(column: str) -> int:
    value = 0
    for char in column.upper():
        if not ("A" <= char <= "Z"):
            raise ValueError(f"非法 Excel 列名：{column}")
        value = value * 26 + ord(char) - ord("A") + 1
    return value


COL = {column: col_to_index(column) for column in set(ITEM_COLUMNS) | set(EXTRA_COLUMN_MAP) | set(BLOCK_COLUMN_MAP.values())}


def parse_yuewei_excel_workbook(file_path: str | Path, sheet_name: str | None = None) -> tuple[dict, list[dict]]:
    """读取工作簿，返回与 excel-imported-blocks.js 兼容的 meta / blocks。

    兼容两类来源：
    1. 早期 Yuewei 成本总表，按固定 A~BE 列位解析。
    2. 钉钉国际物流审批附件中的物料/装箱明细，按表头自动识别解析。
    """

    if load_workbook is None:
        raise RuntimeError("解析 .xlsx 需要安装 openpyxl，请先安装后再导入真实 Excel。")

    path = Path(file_path).expanduser()
    workbook = load_workbook(path, data_only=True, read_only=False)
    try:
        requested_sheet = (sheet_name or "").strip()
        if not requested_sheet and _looks_like_ci_pl_workbook(workbook):
            selected_sheet_name = "CI+PL"
            parser_name = "ci_pl_workbook"
            warning = ""
            blocks = parse_ci_pl_workbook(workbook)
        else:
            selected_sheet_name, parser_name, warning = _select_sheet(workbook, sheet_name)
            worksheet = workbook[selected_sheet_name]
            if parser_name == "oa_attachment_detail":
                blocks = parse_oa_attachment_detail_sheet(worksheet, source_sheet=selected_sheet_name)
            elif parser_name == "sisa_warehouse_receipt":
                blocks = parse_sisa_warehouse_receipt_sheet(worksheet, source_sheet=selected_sheet_name)
            else:
                blocks = parse_yuewei_sheet(worksheet, source_sheet=selected_sheet_name)
        meta = {
            "sheetCount": len(workbook.sheetnames),
            "blockCount": len(blocks),
            "itemCount": sum(len(block.get("items") or []) for block in blocks),
            "sheets": list(workbook.sheetnames),
            "sourceFile": path.name,
            "sourceSheet": selected_sheet_name,
            "requestedSheet": (sheet_name or "").strip(),
            "parser": parser_name,
        }
        if warning:
            meta["warning"] = warning
        return meta, blocks
    finally:
        workbook.close()


def _select_sheet(workbook, sheet_name: str | None) -> tuple[str, str, str]:
    requested_sheet = (sheet_name or "").strip()
    available_sheets = list(workbook.sheetnames)

    if requested_sheet and requested_sheet in workbook.sheetnames:
        parser_name = _detect_sheet_parser(workbook[requested_sheet])
        return requested_sheet, parser_name, ""

    detected_parsers = {"oa_attachment_detail", "sisa_warehouse_receipt"}
    for candidate in available_sheets:
        parser_name = _detect_sheet_parser(workbook[candidate])
        if parser_name in detected_parsers:
            warning = ""
            if requested_sheet:
                warning = f"未找到工作表“{requested_sheet}”，已自动识别并使用“{candidate}”。"
            return candidate, parser_name, warning

    if requested_sheet:
        if len(available_sheets) == 1:
            only_sheet = available_sheets[0]
            parser_name = _detect_sheet_parser(workbook[only_sheet])
            return only_sheet, parser_name, f"未找到工作表“{requested_sheet}”，当前文件只有“{only_sheet}”，已自动使用该工作表。"
        available = "、".join(available_sheets)
        raise ValueError(f"工作簿中不存在工作表：{requested_sheet}。当前文件包含：{available}。也可以把工作表名称留空，由系统自动识别。")

    first_sheet = available_sheets[0]
    return first_sheet, _detect_sheet_parser(workbook[first_sheet]), ""


def _detect_sheet_parser(worksheet) -> str:
    if _find_sisa_warehouse_receipt_header(worksheet):
        return "sisa_warehouse_receipt"
    return "oa_attachment_detail" if _find_oa_attachment_header(worksheet) else "yuewei_cost_workbook"


def parse_yuewei_sheet(worksheet, source_sheet: str | None = None) -> list[dict]:
    source_sheet = source_sheet or worksheet.title
    reader = MergedCellReader(worksheet)
    blocks: list[dict] = []
    current_key = ""
    current_rows: list[int] = []

    def flush_current() -> None:
        nonlocal current_key, current_rows
        if current_rows:
            blocks.append(_build_block(reader, source_sheet, current_rows))
        current_key = ""
        current_rows = []

    for row_no in range(1, worksheet.max_row + 1):
        if not _is_data_row(reader, row_no):
            flush_current()
            continue

        row_key = _block_key(reader, source_sheet, row_no)
        if current_rows and row_key != current_key:
            flush_current()

        current_key = row_key
        current_rows.append(row_no)

    flush_current()
    return blocks


def parse_oa_attachment_detail_sheet(worksheet, source_sheet: str | None = None) -> list[dict]:
    """解析国际物流审批附件里的物料/装箱明细表。

    这类附件通常以“对应钉钉采购订单号、品目编码、总个数、单价、总价、出口方式”
    等中文表头为主，不适合按 A~BE 固定列位读取。
    """

    source_sheet = source_sheet or worksheet.title
    header_row, header_map = _find_oa_attachment_header(worksheet) or (0, {})
    if not header_row:
        return []

    grouped_rows: dict[str, list[dict]] = {}
    row_ranges: dict[str, list[int]] = {}
    for row_no in range(header_row + 1, worksheet.max_row + 1):
        row = _read_attachment_row(worksheet, row_no, header_map)
        if not _is_attachment_data_row(row):
            continue

        group_key = str(row.get("purchase_order_no") or f"{source_sheet}-未关联采购单").strip()
        grouped_rows.setdefault(group_key, []).append(row)
        row_ranges.setdefault(group_key, []).append(row_no)

    blocks = []
    for group_key, rows in grouped_rows.items():
        rows = _normalize_attachment_price_rows(rows)
        row_numbers = row_ranges[group_key]
        first = rows[0]
        block = {
            "id": group_key,
            "batchNo": group_key,
            "sourceSheet": source_sheet,
            "sourceRange": f"{source_sheet}!{row_numbers[0]}:{row_numbers[-1]}",
            "sourceTemplate": "oa_attachment_detail",
            "sourceType": "OA_ATTACHMENT",
            "sourceDocNo": group_key,
            "transportMode": first.get("transport_mode") or _transport_from_sheet_name(source_sheet),
            "projectCollection": first.get("project_collection"),
            "remark": "国际物流审批附件明细",
            "items": [_build_attachment_item(row, source_sheet) for row in rows],
        }
        blocks.append(block)

    return blocks


SISA_WAREHOUSE_HEADER_ALIASES = {
    "image": ("产品图片", "图片"),
    "box_no": ("箱号",),
    "box_count": ("箱数",),
    "material_code": ("型号", "物料编码", "品目编码", "sku", "itemcode"),
    "product_name": ("品名", "物料名称", "产品名称"),
    "declaration_unit_price_usd": ("报关单价",),
    "qty_per_box": ("单箱产品数", "每箱数量"),
    "quantity": ("总产品数", "总个数", "总数量", "数量"),
    "length_cm": ("长cm", "长"),
    "width_cm": ("宽cm", "宽"),
    "height_cm": ("高cm", "高"),
    "volume_per_box_m3": ("体积箱cmb", "体积箱cbm", "体积箱", "体积每箱"),
    "volume_m3": ("总体积cmb", "总体积cbm", "总体积", "体积"),
    "gross_weight_each_kg": ("毛重箱kg", "毛重每箱kg", "毛重箱", "毛重每箱"),
    "gross_weight_kg": ("总毛重kg", "总毛重", "毛重"),
    "source_remark": ("备注",),
}


def parse_sisa_warehouse_receipt_sheet(worksheet, source_sheet: str | None = None) -> list[dict]:
    """解析 SiSA 墨西哥专线进仓单中的“产品清单+派件信息”模板。"""

    source_sheet = source_sheet or worksheet.title
    header_row, header_map = _find_sisa_warehouse_receipt_header(worksheet) or (0, {})
    if not header_row:
        return []

    rows: list[dict] = []
    current_box_no = ""
    current_box_count = None
    for row_no in range(header_row + 1, worksheet.max_row + 1):
        row = _read_sisa_warehouse_receipt_row(
            worksheet,
            row_no=row_no,
            header_map=header_map,
            current_box_no=current_box_no,
            current_box_count=current_box_count,
        )
        if row.get("_stop"):
            break
        if not row:
            continue
        row_box_no = str(row.get("_box_no") or "")
        if row_box_no and not row_box_no.startswith("未标箱号"):
            current_box_no = row_box_no
        if row.get("piece_count") not in (None, ""):
            current_box_count = row.get("piece_count")
        rows.append(row)

    rows = _allocate_sisa_warehouse_physical_fields(rows)
    if not rows:
        return []

    row_numbers = [int(row["excel_row_no"]) for row in rows if row.get("excel_row_no")]
    block_id = _find_sisa_warehouse_batch_id(worksheet, source_sheet)
    block = {
        "id": block_id,
        "batchNo": block_id,
        "sourceSheet": source_sheet,
        "sourceRange": f"{source_sheet}!{min(row_numbers)}:{max(row_numbers)}" if row_numbers else source_sheet,
        "sourceTemplate": "sisa_warehouse_receipt",
        "sourceType": "PACKING_LIST",
        "sourceDocNo": block_id,
        "transportMode": "海运",
        "remark": "SiSA墨西哥专线进仓单产品清单",
        "items": [_build_attachment_item(row, source_sheet) for row in rows],
    }
    return [block]


def _looks_like_ci_pl_workbook(workbook) -> bool:
    sheet_keys = {_normalize_header(name): name for name in workbook.sheetnames}
    return "ci" in sheet_keys and "pl" in sheet_keys


def parse_ci_pl_workbook(workbook) -> list[dict]:
    """解析同一工作簿内的 Commercial Invoice + Packing List。

    CI 里通常有数量、单价、金额；PL 里通常有每箱体积和毛重。
    两者没有物料编码时，按品名中的规格型号合并。
    """

    sheet_keys = {_normalize_header(name): name for name in workbook.sheetnames}
    ci_rows = _read_ci_sheet_rows(workbook[sheet_keys["ci"]], source_sheet=sheet_keys["ci"])
    pl_rows = _read_pl_sheet_rows(workbook[sheet_keys["pl"]], source_sheet=sheet_keys["pl"])
    specs = []
    for row in ci_rows + pl_rows:
        spec = row.get("spec_model")
        if spec and spec not in specs:
            specs.append(spec)

    items = []
    for spec in specs:
        ci_row = next((row for row in ci_rows if row.get("spec_model") == spec), {})
        pl_row = next((row for row in pl_rows if row.get("spec_model") == spec), {})
        product_name = ci_row.get("product_name") or pl_row.get("product_name")
        quantity = ci_row.get("quantity")
        extra = {
            "sourceSheet": "CI+PL",
            "sourceType": "PACKING_LIST",
            "sourceDocNo": ci_row.get("invoice_no") or "",
            "specModel": spec,
            "actualShippedQty": quantity,
            "grossWeightKg": pl_row.get("gross_weight_kg"),
            "volumeM3": pl_row.get("volume_m3"),
            "packageCount": pl_row.get("package_count"),
            "sourceRemark": "CI 数量与 PL 重量体积按规格合并",
        }
        items.append(
            [
                None,
                product_name,
                ci_row.get("unit_price"),
                quantity,
                ci_row.get("goods_value"),
                None,
                None,
                None,
                None,
                None,
                None,
                {key: value for key, value in extra.items() if value not in (None, "")},
            ]
        )

    if not items:
        return []
    invoice_no = next((row.get("invoice_no") for row in ci_rows if row.get("invoice_no")), "")
    return [
        {
            "id": invoice_no or "CI+PL",
            "batchNo": invoice_no or "CI+PL",
            "sourceSheet": "CI+PL",
            "sourceRange": "CI+PL",
            "sourceTemplate": "ci_pl_workbook",
            "sourceType": "PACKING_LIST",
            "sourceDocNo": invoice_no,
            "remark": "商业发票与装箱单合并解析",
            "items": items,
        }
    ]


def _read_ci_sheet_rows(worksheet, source_sheet: str) -> list[dict]:
    header_row, header_map = _find_header_row(
        worksheet,
        required_aliases={
            "article_name": ("Article Name", "品名", "货品名称"),
            "quantity": ("Quantity", "数量"),
        },
        optional_aliases={
            "article_no": ("Article No", "序号"),
            "unit_price": ("Unit Price", "单价"),
            "goods_value": ("Amount", "金额", "总价"),
        },
    )
    if not header_row:
        return []

    invoice_no = _find_ci_invoice_no(worksheet)
    rows = []
    for row_no in range(header_row + 1, worksheet.max_row + 1):
        product_name = _normalize_cell_value(worksheet.cell(row_no, header_map["article_name"]).value)
        if not product_name or "total" in str(product_name).lower():
            continue
        quantity = _normalize_cell_value(worksheet.cell(row_no, header_map["quantity"]).value)
        spec_model = _extract_spec_model(product_name)
        if not spec_model:
            continue
        rows.append(
            {
                "source_sheet": source_sheet,
                "row_no": row_no,
                "invoice_no": invoice_no,
                "product_name": product_name,
                "spec_model": spec_model,
                "quantity": quantity,
                "unit_price": _value_by_header(worksheet, row_no, header_map, "unit_price"),
                "goods_value": _value_by_header(worksheet, row_no, header_map, "goods_value"),
            }
        )
    return rows


def _read_pl_sheet_rows(worksheet, source_sheet: str) -> list[dict]:
    header_row, header_map = _find_header_row(
        worksheet,
        required_aliases={
            "article_name": ("Article Name", "品名", "货品名称"),
            "volume_m3": ("Dimension", "M³", "M3", "CBM", "体积"),
            "gross_weight_kg": ("Weight", "KG", "毛重"),
        },
        optional_aliases={"package_no": ("Package No", "箱号", "包号")},
    )
    if not header_row:
        return []

    grouped: dict[str, dict] = {}
    for row_no in range(header_row + 1, worksheet.max_row + 1):
        product_name = _normalize_cell_value(worksheet.cell(row_no, header_map["article_name"]).value)
        spec_model = _extract_spec_model(product_name)
        if not spec_model:
            continue
        target = grouped.setdefault(
            spec_model,
            {
                "source_sheet": source_sheet,
                "product_name": product_name,
                "spec_model": spec_model,
                "package_count": 0,
                "gross_weight_kg": 0,
                "volume_m3": 0,
            },
        )
        target["package_count"] += 1
        target["gross_weight_kg"] += _to_number(_value_by_header(worksheet, row_no, header_map, "gross_weight_kg"))
        target["volume_m3"] += _to_number(_value_by_header(worksheet, row_no, header_map, "volume_m3"))
    return list(grouped.values())


def _find_header_row(
    worksheet,
    *,
    required_aliases: dict[str, tuple[str, ...]],
    optional_aliases: dict[str, tuple[str, ...]] | None = None,
) -> tuple[int, dict[str, int]]:
    aliases = {**required_aliases, **(optional_aliases or {})}
    for row_no in range(1, min(worksheet.max_row, 40) + 1):
        normalized_headers = {
            col_no: _normalize_header(worksheet.cell(row_no, col_no).value)
            for col_no in range(1, worksheet.max_column + 1)
        }
        field_map: dict[str, int] = {}
        for fieldname, field_aliases in aliases.items():
            for col_no, header in normalized_headers.items():
                if header and any(_normalize_header(alias) in header for alias in field_aliases):
                    field_map[fieldname] = col_no
                    break
        if all(field in field_map for field in required_aliases):
            return row_no, field_map
    return 0, {}


def _find_sisa_warehouse_receipt_header(worksheet) -> tuple[int, dict[str, int]] | None:
    for row_no in range(1, min(worksheet.max_row, 40) + 1):
        normalized_headers = {
            col_no: _normalize_header(worksheet.cell(row_no, col_no).value)
            for col_no in range(1, worksheet.max_column + 1)
        }
        field_map = _build_sisa_warehouse_header_map(normalized_headers)
        if _looks_like_sisa_warehouse_header(worksheet, row_no, field_map):
            return row_no, field_map
    return None


def _build_sisa_warehouse_header_map(normalized_headers: dict[int, str]) -> dict[str, int]:
    field_map: dict[str, int] = {}
    for fieldname, aliases in SISA_WAREHOUSE_HEADER_ALIASES.items():
        for col_no, header in normalized_headers.items():
            if not header:
                continue
            if any(_normalize_header(alias) in header for alias in aliases):
                field_map[fieldname] = col_no
                break
    return field_map


def _looks_like_sisa_warehouse_header(worksheet, row_no: int, field_map: dict[str, int]) -> bool:
    required = {"box_no", "box_count", "material_code", "product_name", "quantity"}
    if not required.issubset(field_map):
        return False

    context_values = [worksheet.title]
    for context_row in range(max(1, row_no - 8), row_no + 1):
        for col_no in range(1, min(worksheet.max_column, 8) + 1):
            value = worksheet.cell(context_row, col_no).value
            if value not in (None, ""):
                context_values.append(str(value))
    context_text = " ".join(context_values)
    return any(keyword in context_text for keyword in ("产品清单", "派件信息", "进仓单", "墨西哥专线", "混装模板"))


def _read_sisa_warehouse_receipt_row(
    worksheet,
    *,
    row_no: int,
    header_map: dict[str, int],
    current_box_no: str,
    current_box_count,
) -> dict:
    values = [_normalize_cell_value(worksheet.cell(row_no, col_no).value) for col_no in range(1, worksheet.max_column + 1)]
    row_text = " ".join(str(value) for value in values if value not in (None, ""))
    if not row_text:
        return {}
    if _normalize_header(values[0]).startswith("total") or "合计" in row_text:
        return {"_stop": True}

    raw_model = _sisa_value(worksheet, row_no, header_map, "material_code")
    raw_product_name = _sisa_value(worksheet, row_no, header_map, "product_name")
    model_is_code = _looks_like_sisa_material_code(raw_model)
    material_code = str(raw_model).strip() if model_is_code else None
    product_name = str(raw_product_name).strip() if raw_product_name not in (None, "") else None
    spec_model = None
    if not model_is_code and raw_model not in (None, ""):
        model_text = str(raw_model).strip()
        if product_name:
            spec_model = model_text
        else:
            product_name = model_text
    if not product_name:
        return {}

    box_no = _sisa_value(worksheet, row_no, header_map, "box_no") or current_box_no
    box_count = _positive_number(_sisa_value(worksheet, row_no, header_map, "box_count"))
    if box_count is None and current_box_count not in (None, ""):
        box_count = _positive_number(current_box_count)
    box_count = box_count or 1

    qty_per_box = _positive_number(_sisa_value(worksheet, row_no, header_map, "qty_per_box"))
    quantity = _positive_number(_sisa_value(worksheet, row_no, header_map, "quantity"))
    if quantity is None and qty_per_box is not None:
        quantity = qty_per_box * box_count

    volume_m3 = _positive_number(_sisa_value(worksheet, row_no, header_map, "volume_m3"))
    if volume_m3 is None:
        volume_per_box = _positive_number(_sisa_value(worksheet, row_no, header_map, "volume_per_box_m3"))
        if volume_per_box is not None:
            volume_m3 = volume_per_box * box_count
    if volume_m3 is None:
        volume_m3 = _calculate_sisa_volume_m3(worksheet, row_no, header_map, box_count)

    gross_weight_kg = _positive_number(_sisa_value(worksheet, row_no, header_map, "gross_weight_kg"))
    if gross_weight_kg is None:
        gross_weight_each = _positive_number(_sisa_value(worksheet, row_no, header_map, "gross_weight_each_kg"))
        if gross_weight_each is not None:
            gross_weight_kg = gross_weight_each * box_count

    declaration_unit_price = _positive_number(_sisa_value(worksheet, row_no, header_map, "declaration_unit_price_usd"))
    source_remark = _join_remark(
        _sisa_value(worksheet, row_no, header_map, "source_remark"),
        f"箱号：{box_no}" if box_no else None,
        f"箱数：{_format_number(box_count)}" if box_count else None,
        f"报关单价USD：{_format_number(declaration_unit_price)}" if declaration_unit_price is not None else None,
    )

    return {
        "excel_row_no": row_no,
        "material_code": material_code,
        "product_name": product_name,
        "spec_model": spec_model,
        "quantity": quantity,
        "qty_per_piece": qty_per_box,
        "piece_count": box_count,
        "gross_weight_kg": gross_weight_kg,
        "volume_m3": volume_m3,
        "transport_mode": "海运",
        "packing": f"{_format_number(box_count)}箱" if box_count else None,
        "source_remark": source_remark,
        "_box_no": box_no or f"未标箱号-{row_no}",
        "_box_total_quantity": quantity,
        "_box_gross_weight_kg": gross_weight_kg,
        "_box_volume_m3": volume_m3,
    }


def _sisa_value(worksheet, row_no: int, header_map: dict[str, int], fieldname: str):
    col_no = header_map.get(fieldname)
    if not col_no:
        return None
    value, _merged_range = _attachment_cell_value(worksheet, row_no, col_no)
    return _normalize_cell_value(value)


def _looks_like_sisa_material_code(value) -> bool:
    if value in (None, ""):
        return False
    text = str(value).strip()
    if not text or text.startswith("="):
        return False
    normalized = _normalize_header(text)
    if normalized in {"total", "合计", "总计"}:
        return False
    if len(text) > 32:
        return False
    if any(char.isspace() for char in text):
        return False
    if any(ord(char) > 127 for char in text):
        return False
    if not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9._/-]*", text):
        return False
    code_with_digits = re.fullmatch(r"[A-Za-z]{1,8}\d{2,}[A-Za-z0-9._/-]*", text)
    uppercase_token = re.fullmatch(r"[A-Z0-9._/-]{2,32}", text) and any("A" <= char <= "Z" for char in text)
    return bool(code_with_digits or uppercase_token)


def _calculate_sisa_volume_m3(worksheet, row_no: int, header_map: dict[str, int], box_count: float) -> float | None:
    length = _positive_number(_sisa_value(worksheet, row_no, header_map, "length_cm"))
    width = _positive_number(_sisa_value(worksheet, row_no, header_map, "width_cm"))
    height = _positive_number(_sisa_value(worksheet, row_no, header_map, "height_cm"))
    if length is None or width is None or height is None:
        return None
    return length * width * height / 1_000_000 * (box_count or 1)


def _allocate_sisa_warehouse_physical_fields(rows: list[dict]) -> list[dict]:
    grouped: dict[str, list[dict]] = {}
    for row in rows:
        grouped.setdefault(str(row.get("_box_no") or row.get("excel_row_no")), []).append(row)

    for group_rows in grouped.values():
        if len(group_rows) <= 1:
            continue

        total_quantity = sum(_positive_number(row.get("quantity")) or 0 for row in group_rows)
        group_weight = next((_positive_number(row.get("_box_gross_weight_kg")) for row in group_rows if _positive_number(row.get("_box_gross_weight_kg"))), None)
        group_volume = next((_positive_number(row.get("_box_volume_m3")) for row in group_rows if _positive_number(row.get("_box_volume_m3"))), None)
        if not total_quantity:
            continue

        for row in group_rows:
            quantity = _positive_number(row.get("quantity")) or 0
            ratio = quantity / total_quantity if total_quantity else 0
            if group_weight is not None:
                row["gross_weight_kg"] = group_weight * ratio
            if group_volume is not None:
                row["volume_m3"] = group_volume * ratio
            if group_weight is not None or group_volume is not None:
                row["source_remark"] = _join_remark(row.get("source_remark"), "同箱重量/体积按产品数量占比分摊")

    return [{key: value for key, value in row.items() if not key.startswith("_")} for row in rows]


def _format_number(value) -> str:
    if value is None:
        return ""
    number = _positive_number(value)
    if number is None:
        return str(value)
    if number.is_integer():
        return str(int(number))
    return f"{number:.6f}".rstrip("0").rstrip(".")


def _find_sisa_warehouse_batch_id(worksheet, source_sheet: str) -> str:
    for row_no in range(1, min(worksheet.max_row, 8) + 1):
        for col_no in range(1, min(worksheet.max_column, 8) + 1):
            header = _normalize_header(worksheet.cell(row_no, col_no).value)
            if header in {"货件号", "客户号"}:
                value = _normalize_cell_value(worksheet.cell(row_no + 1, col_no).value)
                if value not in (None, ""):
                    return str(value).strip()
    return source_sheet


def _value_by_header(worksheet, row_no: int, header_map: dict[str, int], fieldname: str):
    col_no = header_map.get(fieldname)
    if not col_no:
        return None
    return _normalize_cell_value(worksheet.cell(row_no, col_no).value)


def _find_ci_invoice_no(worksheet) -> str:
    for row_no in range(1, min(worksheet.max_row, 12) + 1):
        for col_no in range(1, worksheet.max_column + 1):
            value = _normalize_cell_value(worksheet.cell(row_no, col_no).value)
            text = str(value or "")
            match = re.search(r"C/?I\s*No\.?\s*:?\s*([A-Z0-9-]+)", text, flags=re.IGNORECASE)
            if match:
                return match.group(1).strip()
    return ""


def _extract_spec_model(value) -> str | None:
    text = str(value or "")
    match = re.search(r"\b([A-Z]{1,6}-[A-Z0-9]+(?:-[A-Z0-9]+)*)\b", text, flags=re.IGNORECASE)
    return match.group(1).upper() if match else None


def _to_number(value) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


class MergedCellReader:
    def __init__(self, worksheet):
        self.worksheet = worksheet
        self.max_col = col_to_index(MAX_EXCEL_COLUMN)
        self._merged_values: dict[tuple[int, int], Any] = {}
        for merged_range in worksheet.merged_cells.ranges:
            top_value = _normalize_cell_value(worksheet.cell(merged_range.min_row, merged_range.min_col).value)
            min_col = max(1, merged_range.min_col)
            max_col = min(self.max_col, merged_range.max_col)
            for row_no in range(merged_range.min_row, merged_range.max_row + 1):
                for col_no in range(min_col, max_col + 1):
                    self._merged_values[(row_no, col_no)] = top_value

    def value(self, row_no: int, column: str):
        col_no = COL.get(column) or col_to_index(column)
        value = _normalize_cell_value(self.worksheet.cell(row_no, col_no).value)
        if value not in (None, ""):
            return value
        return self._merged_values.get((row_no, col_no))


def _normalize_cell_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        cleaned = value.strip()
        return cleaned or None
    if isinstance(value, datetime | date | time):
        return value.isoformat()
    return value


def _is_data_row(reader: MergedCellReader, row_no: int) -> bool:
    values = [reader.value(row_no, column) for column in ("A", "B", "C", "D", "E", "F", "G", "H")]
    text = " ".join(str(value) for value in values if value not in (None, ""))
    if not text:
        return False
    if any(marker in text for marker in ("物料编码", "产品名称", "单价", "总货值", "海关进口名称")):
        return False
    return True


def _block_key(reader: MergedCellReader, source_sheet: str, row_no: int) -> str:
    waybill_no = reader.value(row_no, "J")
    customs_no = reader.value(row_no, "I")
    return str(waybill_no or customs_no or f"{source_sheet}!{row_no}").strip()


def _first_row_value(reader: MergedCellReader, rows: list[int], column: str):
    for row_no in rows:
        value = reader.value(row_no, column)
        if value not in (None, ""):
            return value
    return None


def _build_block(reader: MergedCellReader, source_sheet: str, rows: list[int]) -> dict:
    start_row = rows[0]
    end_row = rows[-1]
    block = {
        "id": _first_row_value(reader, rows, "J") or _first_row_value(reader, rows, "I") or f"{source_sheet}!{start_row}",
        "sourceSheet": source_sheet,
        "sourceRange": f"{source_sheet}!{start_row}:{end_row}",
        "items": [_build_item(reader, source_sheet, row_no) for row_no in rows],
    }

    for target_field, column in BLOCK_COLUMN_MAP.items():
        value = _first_row_value(reader, rows, column)
        if value not in (None, ""):
            block[target_field] = value

    if block.get("projectCollection") and not block.get("remark"):
        block["remark"] = block["projectCollection"]
    return block


def _build_item(reader: MergedCellReader, source_sheet: str, row_no: int) -> list:
    item = [reader.value(row_no, column) for column in ITEM_COLUMNS]
    extra = {
        "excelA": reader.value(row_no, "A"),
        "sourceSheet": source_sheet,
        "sourceRow": row_no,
    }

    for column, target_field in EXTRA_COLUMN_MAP.items():
        value = reader.value(row_no, column)
        if value not in (None, ""):
            extra[target_field] = value

    item.append(extra)
    return item


ATTACHMENT_HEADER_ALIASES = {
    "purchase_order_no": ("对应钉钉采购订单号", "采购订单号", "采购订单"),
    "material_code": ("品目编码", "物料编码", "itemcode", "itemno", "partno", "sku"),
    "brand": ("品牌",),
    "import_name": ("申报名称",),
    "unit": ("申报单位", "单位"),
    "supplier": ("供应商", "supplier"),
    "product_name": ("中文品名", "chinesename", "物料名称", "品名"),
    "category": ("物料类别", "大类分类", "大类", "品类"),
    "invoice_flag": ("是否开票", "ci"),
    "hs_code": ("海关编码", "customscode", "hscode"),
    "product_name_en": ("英文品名", "englishproductname"),
    "product_name_es": ("西语品名", "spanishname"),
    "spec_model": ("规格", "型号", "specificationmodelbrand"),
    "length_m": ("长m",),
    "width_m": ("宽m",),
    "height_m": ("高m",),
    "net_weight_each_kg": ("净重nw件kg",),
    "gross_weight_each_kg": ("毛重gw件kg",),
    "unit_cbm": ("单件cbm",),
    "qty_per_piece": ("个数每件", "numberevery"),
    "piece_count": ("件数", "numberofpieces"),
    "quantity": ("总个数", "totalnumberof", "实际发货数量", "发货数量", "数量", "quantity", "qty", "pcs"),
    "packing": ("包装", "packing"),
    "total_net_weight_kg": ("总净重",),
    "gross_weight_kg": ("总毛重", "grossweight", "毛重", "gw"),
    "volume_m3": ("总体积", "totalcapacity", "体积", "volume", "cbm"),
    "volume_weight_kg": ("体积重", "volumeweight"),
    "chargeable_weight_kg": ("计费重", "chargeableweight"),
    "unit_price": ("单价", "unitprice"),
    "goods_value": ("总价", "总金额", "rmb"),
    "planned_ship_date": ("计划出货日期",),
    "source_remark": ("备注", "remarks"),
    "export_mode": ("出口方式",),
    "project_collection": ("项目归属", "项目"),
}


def _find_oa_attachment_header(worksheet) -> tuple[int, dict[str, int]] | None:
    for row_no in range(1, min(worksheet.max_row, 30) + 1):
        normalized_headers = {
            col_no: _normalize_header(worksheet.cell(row_no, col_no).value)
            for col_no in range(1, worksheet.max_column + 1)
        }
        field_map = _build_attachment_header_map(normalized_headers)
        if _looks_like_oa_attachment_header(field_map):
            return row_no, field_map
    return None


def _looks_like_oa_attachment_header(field_map: dict[str, int]) -> bool:
    if "material_code" not in field_map:
        return False

    quantity_hits = sum(1 for field in ("quantity", "piece_count", "qty_per_piece") if field in field_map)
    money_hits = sum(1 for field in ("unit_price", "goods_value") if field in field_map)
    physical_hits = sum(
        1
        for field in (
            "gross_weight_kg",
            "volume_m3",
            "volume_weight_kg",
            "chargeable_weight_kg",
            "total_net_weight_kg",
            "net_weight_each_kg",
            "gross_weight_each_kg",
            "unit_cbm",
        )
        if field in field_map
    )
    source_hits = sum(1 for field in ("purchase_order_no", "export_mode", "project_collection") if field in field_map)
    name_hits = sum(1 for field in ("product_name", "import_name", "product_name_en", "product_name_es", "spec_model") if field in field_map)

    # 采购明细类附件通常有价格/总价；装箱单类附件通常没有价格，
    # 但会有实际发货数量、毛重、体积等物理字段。
    if quantity_hits and money_hits and source_hits:
        return True
    return bool(quantity_hits and (physical_hits or source_hits or name_hits))


def _build_attachment_header_map(normalized_headers: dict[int, str]) -> dict[str, int]:
    field_map: dict[str, int] = {}
    for fieldname, aliases in ATTACHMENT_HEADER_ALIASES.items():
        for alias in aliases:
            normalized_alias = _normalize_header(alias)
            if not normalized_alias:
                continue
            for col_no, header in normalized_headers.items():
                if header and normalized_alias in header:
                    field_map[fieldname] = col_no
                    break
            if fieldname in field_map:
                break
    return field_map


def _normalize_header(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", "").replace("\n", "").replace("\r", "")
    for char in (" ", "\t", "（", "）", "(", ")", "/", "\\", "-", "_", ".", "。", ":", "："):
        text = text.replace(char, "")
    return text.strip().lower()


def _read_attachment_row(worksheet, row_no: int, header_map: dict[str, int]) -> dict:
    row = {"excel_row_no": row_no}
    for fieldname, col_no in header_map.items():
        value, merged_range = _attachment_cell_value(worksheet, row_no, col_no)
        row[fieldname] = _normalize_cell_value(value)
        if fieldname == "unit":
            row[fieldname] = normalize_unit(row[fieldname])
        if merged_range and fieldname in {"unit_price", "goods_value"}:
            row[f"_{fieldname}_merge_range"] = merged_range
    export_mode = row.get("export_mode")
    row["transport_mode"] = _attachment_transport_mode(export_mode, worksheet.title)
    return row


def _attachment_cell_value(worksheet, row_no: int, col_no: int) -> tuple[Any, str]:
    value = worksheet.cell(row_no, col_no).value
    for merged_range in worksheet.merged_cells.ranges:
        if merged_range.min_row <= row_no <= merged_range.max_row and merged_range.min_col <= col_no <= merged_range.max_col:
            top_value = worksheet.cell(merged_range.min_row, merged_range.min_col).value
            return (top_value if top_value not in (None, "") else value), str(merged_range)
    return value, ""


def _positive_number(value) -> float | None:
    number = _to_number(value)
    return number if number > 0 else None


def _almost_same_amount(left: float, right: float) -> bool:
    return abs(left - right) <= max(0.01, abs(right) * 0.000001)


def _attachment_price_group_key(row: dict) -> tuple[str, str]:
    return (
        _normalize_header(row.get("material_code")),
        _normalize_header(row.get("spec_model") or row.get("product_name") or row.get("import_name")),
    )


def _set_attachment_row_price(row: dict, unit_price: float, *, force_goods_value: bool = False) -> None:
    quantity = _positive_number(row.get("quantity") or row.get("qty_per_piece"))
    if not _positive_number(row.get("unit_price")):
        row["unit_price"] = unit_price
    if quantity and (force_goods_value or not _positive_number(row.get("goods_value"))):
        row["goods_value"] = unit_price * quantity


def _normalize_attachment_merged_price_ranges(rows: list[dict]) -> None:
    groups: dict[str, list[dict]] = {}
    for row in rows:
        merge_key = row.get("_unit_price_merge_range") or row.get("_goods_value_merge_range")
        if merge_key:
            groups.setdefault(str(merge_key), []).append(row)

    for group_rows in groups.values():
        if len(group_rows) <= 1:
            continue

        merged_goods_values = [_positive_number(row.get("goods_value")) for row in group_rows]
        merged_goods_values = [value for value in merged_goods_values if value is not None]
        merged_unit_prices = [_positive_number(row.get("unit_price")) for row in group_rows]
        merged_unit_prices = [value for value in merged_unit_prices if value is not None]
        merged_goods_value = merged_goods_values[0] if merged_goods_values else None
        merged_unit_price = merged_unit_prices[0] if merged_unit_prices else None
        total_quantity = sum(_positive_number(row.get("quantity") or row.get("qty_per_piece")) or 0 for row in group_rows)

        if merged_unit_price:
            for row in group_rows:
                _set_attachment_row_price(row, merged_unit_price, force_goods_value=True)
                if _positive_number(row.get("unit_price")) or _positive_number(row.get("goods_value")):
                    row["purchase_currency"] = row.get("purchase_currency") or "人民币RMB"
                    if row.get("_goods_value_merge_range"):
                        row["source_remark"] = _join_remark(
                            row.get("source_remark"),
                            f"合并单价按本行数量重算：源单价 {merged_unit_price}",
                        )
            continue

        if merged_goods_value and total_quantity:
            allocated_unit_price = merged_goods_value / total_quantity
            for row in group_rows:
                quantity = _positive_number(row.get("quantity") or row.get("qty_per_piece"))
                if not quantity:
                    continue
                row["unit_price"] = allocated_unit_price
                row["goods_value"] = allocated_unit_price * quantity
                row["purchase_currency"] = row.get("purchase_currency") or "人民币RMB"
                row["source_remark"] = _join_remark(
                    row.get("source_remark"),
                    f"合并价格按数量拆分：源单价 {merged_unit_price or ''}，源总价 {merged_goods_value}",
                )
            continue


def _normalize_attachment_price_rows(rows: list[dict]) -> list[dict]:
    """把附件里按同物料汇总填写的单价/总价拆回每一行。"""

    normalized_rows = [dict(row) for row in rows]
    _normalize_attachment_merged_price_ranges(normalized_rows)
    groups: dict[tuple[str, str], list[dict]] = {}
    for row in normalized_rows:
        key = _attachment_price_group_key(row)
        if not key[0]:
            continue
        groups.setdefault(key, []).append(row)

    for group_rows in groups.values():
        for row in group_rows:
            unit_price = _positive_number(row.get("unit_price"))
            goods_value = _positive_number(row.get("goods_value"))
            quantity = _positive_number(row.get("quantity") or row.get("qty_per_piece"))
            if not unit_price and goods_value and quantity:
                row["unit_price"] = goods_value / quantity

        unit_prices = [_positive_number(row.get("unit_price")) for row in group_rows]
        unit_prices = [price for price in unit_prices if price is not None]
        if not unit_prices:
            continue

        first_price = unit_prices[0]
        if any(not _almost_same_amount(price, first_price) for price in unit_prices[1:]):
            continue

        total_quantity = sum(_positive_number(row.get("quantity") or row.get("qty_per_piece")) or 0 for row in group_rows)
        goods_values = [_positive_number(row.get("goods_value")) for row in group_rows]
        goods_values = [value for value in goods_values if value is not None]
        split_group_total = (
            len(group_rows) > 1
            and len(goods_values) == 1
            and total_quantity > 0
            and _almost_same_amount(goods_values[0], first_price * total_quantity)
        )

        for row in group_rows:
            _set_attachment_row_price(row, first_price, force_goods_value=split_group_total)
            if _positive_number(row.get("unit_price")) or _positive_number(row.get("goods_value")):
                row["purchase_currency"] = row.get("purchase_currency") or "人民币RMB"

    return normalized_rows


def _is_attachment_data_row(row: dict) -> bool:
    if not _looks_like_material_code(row.get("material_code")):
        return False
    return any(row.get(field) not in (None, "") for field in ("product_name", "import_name", "quantity", "goods_value"))


def _looks_like_material_code(value) -> bool:
    if value in (None, ""):
        return False
    text = str(value).strip()
    if not text:
        return False
    has_ascii_letter = any("A" <= char.upper() <= "Z" for char in text)
    has_digit = any(char.isdigit() for char in text)
    return has_ascii_letter and has_digit


def _transport_from_sheet_name(source_sheet: str) -> str:
    if "空运" in source_sheet:
        return "空运"
    if "快递" in source_sheet:
        return "快递"
    return "海运"


def _attachment_transport_mode(export_mode, source_sheet: str) -> str:
    text = str(export_mode or "").strip()
    if any(keyword in text for keyword in ("海运", "空运", "快递", "express", "Express", "AIR", "Air", "air")):
        return text
    return _transport_from_sheet_name(source_sheet)


def _build_attachment_item(row: dict, source_sheet: str) -> list:
    quantity = row.get("quantity") or row.get("qty_per_piece")
    declared_name = row.get("import_name")
    product_name = row.get("product_name") or declared_name
    source_remark = _join_remark(row.get("source_remark"), f"申报名称：{declared_name}" if declared_name else None)
    extra = {
        "excelA": row.get("material_code"),
        "sourceSheet": source_sheet,
        "sourceRow": row.get("excel_row_no"),
        "sourceType": "OA_ATTACHMENT",
        "sourceDocNo": row.get("purchase_order_no"),
        "purchaseOrderNo": row.get("purchase_order_no"),
        "productNameEs": row.get("product_name_es"),
        "specModel": row.get("spec_model"),
        "unit": normalize_unit(row.get("unit")),
        "actualShippedQty": quantity,
        "grossWeightKg": row.get("gross_weight_kg"),
        "volumeM3": row.get("volume_m3"),
        "volumeWeightKg": row.get("volume_weight_kg"),
        "chargeableWeightKg": row.get("chargeable_weight_kg"),
        "projectCollection": row.get("project_collection"),
        "transportMode": row.get("transport_mode"),
        "purchaseCurrency": row.get("purchase_currency"),
        "sourceRemark": source_remark,
        "supplier": row.get("supplier"),
        "packing": row.get("packing"),
        "plannedShipDate": row.get("planned_ship_date"),
    }
    return [
        row.get("material_code"),
        product_name,
        row.get("unit_price"),
        quantity,
        row.get("goods_value"),
        declared_name,
        row.get("hs_code"),
        row.get("category"),
        None,
        None,
        None,
        {key: value for key, value in extra.items() if value not in (None, "")},
    ]


def _join_remark(*parts) -> str | None:
    cleaned = [str(part).strip() for part in parts if part not in (None, "")]
    return "；".join(cleaned) if cleaned else None

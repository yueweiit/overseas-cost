"""中文用途：批次查询服务测试。"""

from io import BytesIO
import json

from openpyxl import load_workbook

from overseas_costing.services.batch_service import (
    EXCEL_COLUMNS,
    EXTRA_ITEM_FIELDS,
    _build_item_query_args,
    _build_batch_source_status,
    _build_calculation_confirmation_readiness,
    _build_erp_push_payload,
    _build_export_xlsx_content,
    _build_writeback_readiness,
    _build_writeback_field_gaps,
    _export_cell_value,
    _load_erp_push_context,
    _normalize_item_query_filters,
    _normalize_limit,
    _resolve_batch_business_type,
    _resolve_batch_subsidiary_code,
    check_writeback_ready,
    create_batch,
    get_audit_logs,
    get_batch_filter_options,
    get_batch_list,
    get_batch_items,
    is_invalid_approval_status,
    is_hidden_approval_status,
    writeback_to_erp,
)


def test_resolve_batch_subsidiary_code_falls_back_to_saved_dingtalk_entity() -> None:
    batch = {
        "subsidiary_code": "",
        "extra_json": json.dumps(
            {
                "subsidiary": {
                    "subsidiary_code": "YW MOLDES MX模具",
                    "business_entity_name": "YW MOLDES MX模具",
                }
            },
            ensure_ascii=False,
        ),
    }

    assert _resolve_batch_subsidiary_code(batch) == "YW MOLDES MX模具"


def test_resolve_batch_subsidiary_code_keeps_batch_field_first() -> None:
    batch = {
        "subsidiary_code": "Empresas Mexico",
        "extra_json": json.dumps({"subsidiary": {"subsidiary_code": "旧主体"}}, ensure_ascii=False),
    }

    assert _resolve_batch_subsidiary_code(batch) == "Empresas Mexico"


def test_resolve_batch_subsidiary_code_reads_nested_oa_trace_entity() -> None:
    batch = {
        "subsidiary_code": "",
        "extra_json": json.dumps(
            {
                "source": "excel",
                "oa_logistics_trace": {
                    "subsidiary": {"business_entity_name": "产品&开发Departamento de Producto y Desarrollo"}
                },
            },
            ensure_ascii=False,
        ),
    }

    assert _resolve_batch_subsidiary_code(batch) == "产品&开发Departamento de Producto y Desarrollo"


def test_resolve_batch_subsidiary_code_reads_entidad_comercial_form_field() -> None:
    batch = {
        "subsidiary_code": "",
        "extra_json": json.dumps(
            {
                "form_fields": {
                    "业务主体Entidad comercial": {
                        "name": "YW MOLDES/UV",
                    }
                }
            },
            ensure_ascii=False,
        ),
    }

    assert _resolve_batch_subsidiary_code(batch) == "YW MOLDES/UV"


def test_resolve_batch_subsidiary_code_reads_nested_selected_entity_option() -> None:
    batch = {
        "subsidiary_code": "",
        "extra_json": json.dumps(
            {
                "form_fields": {
                    "Entidad comercial": {
                        "selectedOptions": [
                            {"itemId": "ENTITY-001", "name": "Empresas Mexico"}
                        ]
                    }
                }
            },
            ensure_ascii=False,
        ),
    }

    assert _resolve_batch_subsidiary_code(batch) == "Empresas Mexico"


def test_resolve_batch_business_type_falls_back_to_transport_mode() -> None:
    assert _resolve_batch_business_type({"business_type": "", "transport_mode": "SEA"}) == "SEA_STANDARD"
    assert _resolve_batch_business_type(
        {
            "business_type": "",
            "transport_mode": "AIR",
            "extra_json": json.dumps({"business_type": "AIR_DDP"}),
        }
    ) == "AIR_DDP"


def test_normalize_item_query_filters_strips_empty_values() -> None:
    filters = _normalize_item_query_filters(
        {"customs_no": " 26 16 ", "empty": "", "none": None},
        material_code=" YL000098 ",
    )

    assert filters == {
        "customs_no": "26 16",
        "material_code": "YL000098",
    }


def test_build_item_query_args_adds_field_and_keyword_filters() -> None:
    db_filters, or_filters = _build_item_query_args(
        "BATCH-DOC",
        "VERSION-DOC",
        {"customs_no": "6000151", "keyword": "TPU"},
    )

    assert ["batch", "=", "BATCH-DOC"] in db_filters
    assert ["version", "=", "VERSION-DOC"] in db_filters
    assert ["customs_no", "like", "%6000151%"] in db_filters
    assert ["product_name", "like", "%TPU%"] in or_filters
    assert ["material_code", "like", "%TPU%"] in or_filters


def test_get_batch_items_dry_run_keeps_filters() -> None:
    result = get_batch_items(
        batch_name="HPCU5155607",
        product_name="TPU",
        keyword="YL000098",
    )

    assert result["ok"] is True
    assert result["dry_run"] is True
    assert result["filters"]["product_name"] == "TPU"
    assert result["filters"]["keyword"] == "YL000098"
    assert result["columns"][0]["fieldname"] == "material_code"
    assert "derived_json" in EXTRA_ITEM_FIELDS


def test_excel_columns_include_spec_model_after_product_name() -> None:
    fieldnames = [column["fieldname"] for column in EXCEL_COLUMNS]
    spec_column = next(column for column in EXCEL_COLUMNS if column["fieldname"] == "spec_model")
    quantity_column = next(column for column in EXCEL_COLUMNS if column["fieldname"] == "quantity")
    shipped_quantity_column = next(column for column in EXCEL_COLUMNS if column["fieldname"] == "actual_shipped_qty")
    unit_column = next(column for column in EXCEL_COLUMNS if column["fieldname"] == "unit")
    total_unit_column = next(column for column in EXCEL_COLUMNS if column["fieldname"] == "total_unit_rmb")

    assert fieldnames.index("spec_model") == fieldnames.index("product_name") + 1
    assert fieldnames.index("actual_shipped_qty") == fieldnames.index("quantity") + 1
    assert fieldnames.index("unit") == fieldnames.index("actual_shipped_qty") + 1
    assert "Especificación / Modelo" in spec_column["label"]
    assert quantity_column["label"] == "采购数量"
    assert shipped_quantity_column["label"] == "出库数量（实际发货）"
    assert unit_column["label"] == "单位"
    assert total_unit_column["label"] == "综合单价 RMB"
    assert "business_type" not in fieldnames


def test_sea_air_express_share_same_item_columns() -> None:
    fieldnames = [column["fieldname"] for column in EXCEL_COLUMNS]

    assert fieldnames[:9] == [
        "material_code",
        "product_name",
        "spec_model",
        "unit_price",
        "purchase_currency",
        "quantity",
        "actual_shipped_qty",
        "unit",
        "goods_value",
    ]
    assert fieldnames == [column["fieldname"] for column in EXCEL_COLUMNS]
    assert {"spec_model", "product_name_es", "unit", "purchase_currency"}.issubset(set(EXTRA_ITEM_FIELDS + fieldnames))


def test_get_batch_list_defaults_to_recent_days_with_classic_samples(monkeypatch) -> None:
    from overseas_costing.services import batch_service

    calls = []

    class FakeFrappe:
        @staticmethod
        def get_all(doctype, **kwargs):
            calls.append((doctype, kwargs))
            if kwargs.get("or_filters"):
                return [
                    {
                        "name": "BATCH-HPCU",
                        "batch_no": "HPCU5155607",
                        "transport_mode": "SEA",
                        "source_created_at": "2026-01-29 10:20:00",
                        "source_approval_status": "COMPLETED",
                        "extra_json": "{}",
                        "modified": "2026-08-10 10:00:00",
                    }
                ]
            return [
                {
                    "name": "BATCH-RECENT",
                    "batch_no": "202608051608000144099",
                    "transport_mode": "SEA",
                    "source_created_at": "2026-08-05 16:08:00",
                    "source_approval_status": "RUNNING",
                    "extra_json": "{}",
                    "modified": "2026-08-05 16:08:00",
                }
            ]

    monkeypatch.setattr(batch_service, "frappe", FakeFrappe)
    monkeypatch.setattr(batch_service, "_recent_start", lambda recent_days: "2026-07-15 00:00:00")
    monkeypatch.setattr(batch_service, "_attach_batch_source_status", lambda items: items)
    monkeypatch.setattr(batch_service, "_attach_batch_calculation_snapshot", lambda items: items)

    result = get_batch_list({"transport_mode": "", "recent_days": 30})

    assert result["ok"] is True
    assert result["total"] == 2
    assert len(calls) == 2
    assert calls[0][1]["filters"] == [["source_created_at", ">=", "2026-07-15 00:00:00"]]
    assert calls[1][1]["filters"] == []
    assert ["batch_no", "=", "HPCU5155607"] in calls[1][1]["or_filters"]
    classic_item = next(item for item in result["items"] if item["batch_no"] == "HPCU5155607")
    assert classic_item["is_classic_sample"] == 1
    assert classic_item["sample_note"]


def test_get_batch_list_keeps_invalid_historical_approval_for_traceability(monkeypatch) -> None:
    from overseas_costing.services import batch_service

    class FakeFrappe:
        @staticmethod
        def get_all(doctype, **kwargs):
            assert doctype == "Overseas Cost Batch"
            return [
                {
                    "name": "BATCH-TERMINATED",
                    "batch_no": "202608250001",
                    "transport_mode": "SEA",
                    "source_created_at": "2026-08-25 10:00:00",
                    "source_approval_status": "TERMINATED",
                    "extra_json": "{}",
                    "modified": "2026-08-25 10:00:00",
                }
            ]

    def attach_source_status(items):
        for item in items:
            item["source_status"] = {
                "invalid_business": True,
                "invalid_business_scope": "source_approval",
            }
        return items

    monkeypatch.setattr(batch_service, "frappe", FakeFrappe)
    monkeypatch.setattr(batch_service, "_attach_batch_source_status", attach_source_status)
    monkeypatch.setattr(batch_service, "_attach_batch_calculation_snapshot", lambda items: items)

    result = get_batch_list({"transport_mode": "", "recent_days": 30, "include_history": 1})

    assert result["total"] == 1
    assert result["items"][0]["name"] == "BATCH-TERMINATED"
    assert result["items"][0]["source_status"]["invalid_business"] is True


def test_get_batch_list_treats_all_transport_mode_as_no_filter(monkeypatch) -> None:
    from overseas_costing.services import batch_service

    calls = []

    class FakeFrappe:
        @staticmethod
        def get_all(doctype, **kwargs):
            calls.append((doctype, kwargs))
            return []

    monkeypatch.setattr(batch_service, "frappe", FakeFrappe)
    monkeypatch.setattr(batch_service, "_attach_batch_source_status", lambda items: items)
    monkeypatch.setattr(batch_service, "_attach_batch_calculation_snapshot", lambda items: items)
    monkeypatch.setattr(batch_service, "_recent_start", lambda recent_days: "2026-07-15 00:00:00")

    result = get_batch_list({"transport_mode": "ALL", "recent_days": 30, "include_history": 1})

    assert result["ok"] is True
    assert calls[0][0] == "Overseas Cost Batch"
    assert ["transport_mode", "=", "ALL"] not in calls[0][1]["filters"]
    assert not any(field_filter[0] == "transport_mode" for field_filter in calls[0][1]["filters"])


def test_get_batch_list_keyword_can_find_history_batch_by_item_field(monkeypatch) -> None:
    from overseas_costing.services import batch_service

    calls = []

    class FakeFrappe:
        @staticmethod
        def get_all(doctype, **kwargs):
            calls.append((doctype, kwargs))
            if doctype == "Overseas Cost Item":
                return [{"batch": "HISTORY-BATCH"}]
            if doctype == "Overseas Cost Batch" and kwargs.get("filters") and ["name", "in", ["HISTORY-BATCH"]] in kwargs["filters"]:
                return [
                    {
                        "name": "HISTORY-BATCH",
                        "batch_no": "202606010001",
                        "transport_mode": "SEA",
                        "source_created_at": "2026-06-01 10:00:00",
                        "source_approval_status": "COMPLETED",
                        "extra_json": "{}",
                        "modified": "2026-06-01 10:00:00",
                    }
                ]
            return []

    monkeypatch.setattr(batch_service, "frappe", FakeFrappe)
    monkeypatch.setattr(batch_service, "_attach_batch_source_status", lambda items: items)
    monkeypatch.setattr(batch_service, "_attach_batch_calculation_snapshot", lambda items: items)

    result = get_batch_list({"transport_mode": "", "recent_days": 30, "keyword": "墨镜"})

    assert result["ok"] is True
    assert result["total"] == 1
    assert result["items"][0]["name"] == "HISTORY-BATCH"
    assert any(call[0] == "Overseas Cost Item" for call in calls)


def test_get_batch_list_uses_explicit_source_created_date_range(monkeypatch) -> None:
    from overseas_costing.services import batch_service

    calls = []

    class FakeFrappe:
        @staticmethod
        def get_all(doctype, **kwargs):
            calls.append((doctype, kwargs))
            return []

    monkeypatch.setattr(batch_service, "frappe", FakeFrappe)
    monkeypatch.setattr(batch_service, "_attach_batch_source_status", lambda items: items)
    monkeypatch.setattr(batch_service, "_attach_batch_calculation_snapshot", lambda items: items)

    result = get_batch_list(
        {
            "transport_mode": "",
            "recent_days": 30,
            "start_date": "2026-07-21",
            "end_date": "2026-08-21",
        }
    )

    assert result["ok"] is True
    assert calls[0][1]["filters"] == [
        ["source_created_at", ">=", "2026-07-21 00:00:00"],
        ["source_created_at", "<=", "2026-08-21 23:59:59"],
    ]
    assert "or_filters" not in calls[0][1]


def test_get_batch_filter_options_uses_all_dingtalk_entities(monkeypatch) -> None:
    from overseas_costing.services import batch_service

    class FakeFrappe:
        @staticmethod
        def get_all(doctype, **kwargs):
            assert doctype == "Overseas Cost Batch"
            assert kwargs["filters"] == {}
            return [
                {"subsidiary_code": "LEMOS MX", "extra_json": "{}"},
                {"subsidiary_code": "", "extra_json": json.dumps({"subsidiary": {"name": "YW MOLDES/UV"}})},
                {"subsidiary_code": "LEMOS MX", "extra_json": "{}"},
            ]

    monkeypatch.setattr(batch_service, "frappe", FakeFrappe)

    result = get_batch_filter_options()

    assert result["ok"] is True
    assert result["items"] == ["LEMOS MX", "YW MOLDES/UV"]


def test_create_batch_dry_run_builds_manual_batch_and_version() -> None:
    result = create_batch(
        json.dumps(
            {
                "batch_no": "MANUAL-001",
                "customs_no": "26 16 1681 6000151",
                "waybill_no": "HPCU5155607",
                "transport_mode": "海运",
                "business_type": "SEA_DDP",
                "project_collection": "Yuewei",
                "source_dingtalk_url": "https://aflow.dingtalk.com/dingtalk/web/query/pchomepage.htm#/plainapproval?procInstId=PROC-MANUAL-001",
            },
            ensure_ascii=False,
        )
    )

    assert result["ok"] is True
    assert result["dry_run"] is True
    assert result["batch"]["batch_no"] == "MANUAL-001"
    assert result["batch"]["waybill_no"] == "HPCU5155607"
    assert result["batch"]["transport_mode"] == "SEA"
    assert result["batch"]["business_type"] == "SEA_DDP"
    assert result["batch"]["source_type"] == "manual"
    assert result["batch"]["source_instance_id"] == "PROC-MANUAL-001"
    assert result["version"]["version_code"] == "手工-MANUAL-001"
    assert result["version"]["is_current"] == 1


def test_create_batch_dry_run_requires_batch_no() -> None:
    result = create_batch({"transport_mode": "空运"})

    assert result["ok"] is False
    assert result["dry_run"] is True
    assert "批次号" in result["message"]


def test_normalize_limit_keeps_audit_queries_bounded() -> None:
    assert _normalize_limit("20") == 20
    assert _normalize_limit("bad") == 80
    assert _normalize_limit(999) == 300
    assert _normalize_limit(0) == 1


def test_get_audit_logs_dry_run_returns_stable_shape() -> None:
    result = get_audit_logs(batch_name="HPCU5155607", version_name="VERSION-001", limit=20)

    assert result["ok"] is True
    assert result["dry_run"] is True
    assert result["batch_name"] == "HPCU5155607"
    assert result["version_name"] == "VERSION-001"
    assert result["items"] == []
    assert result["total"] == 0


def test_build_export_xlsx_content_styles_and_freezes_header() -> None:
    content = _build_export_xlsx_content(
        columns=[
            {"excel_col": "A", "fieldname": "material_code", "label": "物料编码"},
            {"excel_col": "B", "fieldname": "product_name", "label": "产品名称"},
        ],
        rows=[["FL004106", "钢化膜"]],
    )

    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active

    assert sheet.freeze_panes == "C2"
    assert sheet.auto_filter.ref == "A1:B2"
    assert sheet["A1"].value == "物料编码"
    assert sheet["A1"].fill.fgColor.rgb == "FFFFD966"
    assert sheet["A1"].font.sz == 8
    assert sheet["A1"].font.bold is False
    assert sheet["A1"].font.color.rgb == "FF000000"
    assert sheet["A1"].border.left.style == "thin"
    assert sheet["A1"].border.left.color.rgb == "FF000000"
    assert sheet["A2"].value == "FL004106"


def test_export_cell_value_hides_default_zero_cleaning_fee() -> None:
    column = {"fieldname": "limpieza_contenedor"}

    assert _export_cell_value({"limpieza_contenedor": 0}, {}, column) == ""
    assert _export_cell_value({"limpieza_contenedor": "0.00"}, {}, column) == ""
    assert _export_cell_value({"limpieza_contenedor": 25}, {}, column) == 25


def test_build_export_xlsx_content_merges_repeated_batch_level_cells() -> None:
    content = _build_export_xlsx_content(
        columns=[
            {"excel_col": "A", "fieldname": "material_code", "label": "物料编码"},
            {"excel_col": "I", "fieldname": "customs_no", "label": "报关单号"},
            {"excel_col": "J", "fieldname": "waybill_no", "label": "中国到墨西哥运单号"},
            {"excel_col": "K", "fieldname": "china_misc_rmb", "label": "中国运输及相关杂费 RMB"},
        ],
        rows=[
            ["FL001", "26 16 1681 6000151", "HPCU5155607", 0],
            ["FL001", "26 16 1681 6000151", "HPCU5155607", 0],
            ["FL002", "26 16 1681 6000151", "FSCU8486789", 0],
        ],
    )

    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active
    merged_ranges = {str(cell_range) for cell_range in sheet.merged_cells.ranges}

    assert "A2:A3" not in merged_ranges
    assert "B2:B4" in merged_ranges
    assert "C2:C3" in merged_ranges
    assert "D2:D4" in merged_ranges
    assert sheet["B2"].alignment.horizontal == "center"
    assert sheet["B2"].alignment.vertical == "center"


def test_build_export_xlsx_content_expands_long_text_columns() -> None:
    content = _build_export_xlsx_content(
        columns=[
            {"excel_col": "BD", "fieldname": "project_collection", "label": "项目归集"},
            {"excel_col": "BE", "fieldname": "transport_mode", "label": "运输方式"},
        ],
        rows=[
            ["productos comerciales贸易产品", "海运"],
            ["YW OEM-Tablet平板", "海运"],
        ],
    )

    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active

    assert sheet.column_dimensions["A"].width >= 36
    assert sheet.column_dimensions["B"].width >= 14
    assert sheet["A2"].alignment.wrap_text is True
    assert sheet["B2"].alignment.wrap_text is True
    assert sheet.row_dimensions[2].height >= 20


def test_build_export_xlsx_content_uses_reference_header_color_groups() -> None:
    content = _build_export_xlsx_content(
        columns=[
            {"excel_col": "N", "fieldname": "cc_rate", "label": "C.C税率"},
            {"excel_col": "X", "fieldname": "import_tax_total", "label": "IMPUESTOS合计清关税费"},
            {"excel_col": "AP", "fieldname": "mexico_customs_rmb", "label": "墨西哥清关费用 RMB"},
            {"excel_col": "AX", "fieldname": "freight_alloc_rmb", "label": "运输费用分摊 RMB"},
            {"excel_col": "BB", "fieldname": "total_cost_rmb", "label": "综合成本 RMB"},
        ],
        rows=[[0, 0, 0, 0, 0]],
    )

    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active

    assert sheet["A1"].fill.fgColor.rgb == "FF0DF2DF"
    assert sheet["B1"].fill.fgColor.rgb == "FFD04FCA"
    assert sheet["C1"].fill.fgColor.rgb == "FFFCC102"
    assert sheet["D1"].fill.fgColor.rgb == "FFFFF3CE"
    assert sheet["E1"].fill.fgColor.rgb == "FF04B0F1"
    assert sheet["E1"].font.bold is True
    assert sheet["E1"].font.color.rgb == "FFFFFFFF"


def test_check_writeback_ready_dry_run_returns_blocking_reasons() -> None:
    result = check_writeback_ready(batch_name="HPCU5155607", version_name="VERSION-001")

    assert result["ok"] is True
    assert result["dry_run"] is True
    assert result["ready"] is False
    assert result["checks"]["batch_exists"] is False
    assert result["blocking_reasons"] == ["当前未连接 Frappe，不能执行真实回写检查。"]
    assert result["item_issue_examples"] == []


def test_load_erp_push_context_skips_subsidiary_column_when_schema_lags(monkeypatch) -> None:
    from overseas_costing.services import batch_service

    queried_fields = []

    class FakeDB:
        @staticmethod
        def has_column(doctype, fieldname):
            assert doctype == "Overseas Cost Batch"
            assert fieldname == "subsidiary_code"
            return False

        @staticmethod
        def get_value(doctype, name, fields, as_dict=False):
            queried_fields.append((doctype, fields))
            if doctype == "Overseas Cost Batch":
                assert "subsidiary_code" not in fields
                return {
                    "name": name,
                    "batch_no": "BATCH-001",
                    "status": "Calculated",
                    "confirm_status": "Pending",
                    "current_version": "VERSION-001",
                    "item_count": 0,
                    "estimated_total_cost_rmb": 0,
                    "actual_total_cost_rmb": 0,
                }
            return {"name": name, "rule_snapshot_json": "[]"}

    class FakeFrappe:
        db = FakeDB()

        @staticmethod
        def get_all(*args, **kwargs):
            return []

    monkeypatch.setattr(batch_service, "frappe", FakeFrappe)
    monkeypatch.setattr(batch_service, "_resolve_batch_name", lambda batch_name: "BATCH-DOC")
    monkeypatch.setattr(batch_service, "_resolve_version_name", lambda batch_name, version_name=None: "VERSION-DOC")

    result = _load_erp_push_context("BATCH-DOC", "VERSION-DOC")

    assert result["ok"] is True
    assert result["batch"]["subsidiary_code"] == ""
    assert queried_fields[0][0] == "Overseas Cost Batch"


def test_build_writeback_readiness_allows_complete_confirmed_batch() -> None:
    result = _build_writeback_readiness(
        batch={
            "status": "Clean",
            "confirm_status": "Confirmed",
            "current_version": "VERSION-001",
            "subsidiary_code": "MX01",
            "item_count": 1,
            "actual_total_cost_rmb": 25,
        },
        resolved_version_name="VERSION-001",
        items=[
            {
                "row_no": 1,
                "material_code": "YL000001",
                "product_name": "太阳眼镜",
                "quantity": 2,
                "actual_shipped_qty": 2,
                "unit_price": 8,
                "purchase_currency": "RMB",
                "goods_value": 16,
                "total_unit_rmb": 12.5,
            }
        ],
    )

    assert result["ready"] is True
    assert result["blocking_reasons"] == []
    assert result["checks"]["has_items"] is True
    assert result["checks"]["items_have_unit_price"] is True


def test_build_writeback_readiness_blocks_incomplete_item_data() -> None:
    result = _build_writeback_readiness(
        batch={
            "status": "Dirty",
            "confirm_status": "Draft",
            "current_version": "",
            "subsidiary_code": "",
            "item_count": 2,
            "estimated_total_cost_rmb": 0,
            "actual_total_cost_rmb": 0,
        },
        resolved_version_name=None,
        items=[
            {
                "row_no": 7,
                "material_code": "",
                "product_name": "保护膜",
                "quantity": 0,
                "actual_shipped_qty": 0,
                "unit_price": "",
                "purchase_currency": "",
                "goods_value": 0,
                "total_unit_rmb": 0,
            }
        ],
    )

    assert result["ready"] is False
    assert result["checks"]["has_current_version"] is False
    assert result["checks"]["has_dirty_data"] is True
    assert result["item_issue_counts"]["material_code"] == 1
    assert result["item_issue_counts"]["actual_shipped_qty"] == 1
    assert result["item_issue_counts"]["unit_price"] == 1
    assert result["item_issue_examples"][0]["row_no"] == 7
    assert "当前批次还没有确认。" in result["blocking_reasons"]
    assert "批次记录明细数为 2，实际查询到 1 条。" in result["warning_reasons"]


def test_rejected_linked_purchase_approval_blocks_confirmation_and_writeback() -> None:
    batch = {
        "status": "Calculated",
        "confirm_status": "Confirmed",
        "current_version": "VERSION-001",
        "subsidiary_code": "MX01",
        "item_count": 1,
        "actual_total_cost_rmb": 25,
        "extra_json": json.dumps(
            {
                "linked_purchase_approvals": [
                    {
                        "approval_no": "PUR-REJECTED-001",
                        "approval_status": "REJECTED",
                        "message": "总经理拒绝，本次不采购。",
                    }
                ]
            },
            ensure_ascii=False,
        ),
    }
    items = [
        {
            "row_no": 1,
            "material_code": "YL000001",
            "product_name": "太阳眼镜",
            "quantity": 2,
            "actual_shipped_qty": 2,
            "unit_price": 8,
            "purchase_currency": "RMB",
            "goods_value": 16,
            "freight_alloc_rmb": 6,
            "mexico_customs_mxn": 10,
            "import_tax_total": 4,
            "total_unit_rmb": 12.5,
        }
    ]
    rules = [
        {"expense_category": "国际运费", "amount": 6},
        {"expense_category": "清关费", "amount": 10},
        {"expense_category": "关税", "amount": 4},
    ]

    confirmation = _build_calculation_confirmation_readiness(
        batch=batch,
        items=items,
        rules=rules,
        resolved_version_name="VERSION-001",
    )
    writeback = _build_writeback_readiness(
        batch=batch,
        items=items,
        rules=rules,
        resolved_version_name="VERSION-001",
    )

    assert confirmation["ready"] is False
    assert writeback["ready"] is False
    assert confirmation["checks"]["has_invalid_business_approval"] is True
    assert writeback["checks"]["has_invalid_business_approval"] is True
    assert "PUR-REJECTED-001" in confirmation["blocking_reasons"][0]
    assert "不进入综合成本确认或 ERP 推送" in writeback["blocking_reasons"][0]


def test_build_batch_source_status_exposes_invalid_linked_purchase_approval() -> None:
    status = _build_batch_source_status(
        {
            "name": "BATCH-REJECTED",
            "batch_no": "202608181457000111961",
            "source_type": "oa_logistics",
            "source_approval_no": "LOGISTICS-001",
            "extra_json": json.dumps(
                {
                    "linked_purchase_approvals": [
                        {
                            "approval_no": "PUR-REJECTED-001",
                            "source_instance_id": "PROC-PUR-001",
                            "approval_status": "REJECTED",
                            "message": "总经理拒绝，本次不采购。",
                        }
                    ]
                },
                ensure_ascii=False,
            ),
        }
    )

    assert status["invalid_business"] is True
    assert status["invalid_business_scope"] == "linked_purchase_approval"
    assert status["purchase_approval_sync_state"] == "invalid"
    assert status["linked_purchase_count"] == 1
    assert status["invalid_purchase_approval_count"] == 1
    assert status["linked_purchase_approval_nos"] == ["PUR-REJECTED-001"]


def test_build_batch_source_status_explains_missing_and_pending_purchase_approval() -> None:
    missing = _build_batch_source_status(
        {
            "name": "BATCH-MISSING",
            "source_type": "oa_logistics",
            "source_approval_no": "LOGISTICS-001",
        }
    )
    pending = _build_batch_source_status(
        {
            "name": "BATCH-PENDING",
            "source_type": "oa_logistics",
            "source_approval_no": "LOGISTICS-002",
            "extra_json": json.dumps(
                {"linked_purchase_approvals": [{"approval_no": "PUR-PENDING-001"}]},
                ensure_ascii=False,
            ),
        }
    )

    assert missing["invalid_business"] is False
    assert missing["purchase_approval_sync_state"] == "missing"
    assert "未关联采购审批" in missing["purchase_approval_sync_message"]
    assert pending["invalid_business"] is False
    assert pending["purchase_approval_sync_state"] == "pending"
    assert "尚未同步" in pending["purchase_approval_sync_message"]


def test_build_calculation_confirmation_readiness_requires_subsidiary_and_fee_pools() -> None:
    result = _build_calculation_confirmation_readiness(
        batch={
            "status": "Calculated",
            "confirm_status": "Pending",
            "current_version": "VERSION-001",
            "subsidiary_code": "",
            "item_count": 1,
            "estimated_total_cost_rmb": 25,
        },
        resolved_version_name="VERSION-001",
        rules=[],
        items=[
            {
                "row_no": 1,
                "material_code": "YL000001",
                "product_name": "太阳眼镜",
                "quantity": 2,
                "actual_shipped_qty": 2,
                "unit_price": 8,
                "purchase_currency": "RMB",
                "goods_value": 16,
                "total_unit_rmb": 12.5,
            }
        ],
    )

    assert result["ready"] is False
    assert result["checks"]["has_subsidiary_code"] is False
    assert result["checks"]["has_international_freight"] is False
    assert "当前批次缺少归属业务主体。" in result["blocking_reasons"]
    assert "当前批次缺少国际运费费用池或分摊结果。" in result["blocking_reasons"]
    assert any(gap["fieldname"] == "subsidiary_code" for gap in result["field_gaps"]["batch"])
    assert any(gap["fieldname"] == "国际运费" for gap in result["field_gaps"]["rules"])


def test_build_calculation_confirmation_readiness_uses_saved_entity_fallback() -> None:
    result = _build_calculation_confirmation_readiness(
        batch={
            "subsidiary_code": "",
            "extra_json": json.dumps({"subsidiary": {"name": "Empresas Mexico"}}, ensure_ascii=False),
            "status": "Calculated",
            "current_version": "VERSION-001",
            "item_count": 1,
            "estimated_total_cost_rmb": 20,
        },
        items=[
            {
                "material_code": "ITEM-001",
                "product_name": "物料",
                "quantity": 2,
                "actual_shipped_qty": 2,
                "unit_price": 8,
                "purchase_currency": "RMB",
                "goods_value": 16,
                "freight_alloc_rmb": 4,
                "mexico_customs_mxn": 1,
                "import_tax_total": 1,
                "total_unit_rmb": 10,
            }
        ],
        rules=[],
        resolved_version_name="VERSION-001",
    )

    assert result["checks"]["has_subsidiary_code"] is True


def test_build_calculation_confirmation_readiness_allows_complete_calculation() -> None:
    result = _build_calculation_confirmation_readiness(
        batch={
            "status": "Calculated",
            "confirm_status": "Pending",
            "current_version": "VERSION-001",
            "subsidiary_code": "MX01",
            "item_count": 1,
            "estimated_total_cost_rmb": 25,
        },
        resolved_version_name="VERSION-001",
        rules=[
            {
                "rule_code": "china_ocean_usd",
                "expense_category": "中国海运费",
                "allocation_basis": "gross_weight",
                "currency": "USD",
                "amount": 30,
            },
            {
                "rule_code": "mexico_customs_mxn",
                "expense_category": "墨西哥清关费",
                "allocation_basis": "goods_value",
                "currency": "MXN",
                "amount": 10,
            },
            {
                "rule_code": "import_tax_total",
                "expense_category": "关税税费",
                "allocation_basis": "goods_value",
                "currency": "MXN",
                "amount": 4,
            },
        ],
        items=[
            {
                "row_no": 1,
                "material_code": "YL000001",
                "product_name": "太阳眼镜",
                "quantity": 2,
                "actual_shipped_qty": 2,
                "unit_price": 8,
                "purchase_currency": "RMB",
                "goods_value": 16,
                "freight_alloc_rmb": 6,
                "mexico_customs_mxn": 10,
                "import_tax_total": 4,
                "total_unit_rmb": 12.5,
            }
        ],
    )

    assert result["ready"] is True
    assert result["checks"]["has_subsidiary_code"] is True
    assert result["checks"]["has_clearance_fee"] is True
    assert result["checks"]["has_tariff"] is True


def test_calculation_readiness_accepts_zero_confirmed_clearance_and_tariff() -> None:
    result = _build_calculation_confirmation_readiness(
        batch={
            "status": "Calculated",
            "confirm_status": "Pending",
            "current_version": "VERSION-001",
            "subsidiary_code": "Empresas Dragon",
            "item_count": 1,
            "estimated_total_cost_rmb": 16,
        },
        resolved_version_name="VERSION-001",
        rules=[
            {
                "rule_code": "china_ocean_usd",
                "expense_category": "中国海运费",
                "allocation_basis": "gross_weight",
                "currency": "USD",
                "amount": 30,
            },
            {
                "rule_code": "manual_clearance_fee",
                "expense_category": "清关费",
                "allocation_basis": "gross_weight",
                "currency": "MXN",
                "amount": 0,
                "remark": "OCW_ZERO_CONFIRMED | 人工确认本票清关费为0",
            },
            {
                "rule_code": "manual_tariff_tax",
                "expense_category": "关税税费",
                "allocation_basis": "goods_value",
                "currency": "MXN",
                "amount": 0,
                "remark": "OCW_ZERO_CONFIRMED | 人工确认本票关税为0",
            },
        ],
        items=[
            {
                "row_no": 1,
                "material_code": "YL000001",
                "product_name": "太阳眼镜",
                "quantity": 2,
                "actual_shipped_qty": 2,
                "unit_price": 8,
                "purchase_currency": "RMB",
                "goods_value": 16,
                "freight_alloc_rmb": 6,
                "total_unit_rmb": 8,
            }
        ],
    )

    assert result["ready"] is True
    assert result["checks"]["has_clearance_fee"] is True
    assert result["checks"]["has_tariff"] is True
    assert not any(gap["fieldname"] == "清关费" for gap in result["field_gaps"]["rules"])
    assert not any(gap["fieldname"] == "关税" for gap in result["field_gaps"]["rules"])


def test_build_writeback_field_gaps_groups_missing_info_by_scope() -> None:
    gaps = _build_writeback_field_gaps(
        batch={
            "subsidiary_code": "",
            "current_version": "",
            "estimated_total_cost_rmb": 0,
            "actual_total_cost_rmb": 0,
        },
        items=[
            {
                "row_no": 1,
                "material_code": "",
                "product_name": "太阳眼镜",
                "quantity": 0,
                "actual_shipped_qty": 0,
                "unit_price": 0,
                "purchase_currency": "",
                "goods_value": 0,
                "total_unit_rmb": 0,
            }
        ],
        rules=[],
        resolved_version_name=None,
    )

    assert any(gap["fieldname"] == "subsidiary_code" for gap in gaps["batch"])
    assert any(gap["fieldname"] == "current_version" for gap in gaps["batch"])
    assert any(gap["fieldname"] == "material_code" for gap in gaps["items"])
    assert any(gap["fieldname"] == "国际运费" for gap in gaps["rules"])


def test_build_erp_push_payload_contains_core_fields_and_expense_details() -> None:
    payload = _build_erp_push_payload(
        batch={"name": "BATCH-001", "batch_no": "HPCU5155607", "subsidiary_code": "MX01"},
        version={"name": "VERSION-001", "version_code": "V1"},
        readiness={"total_cost_rmb": 25},
        rules=[
            {
                "rule_code": "china_ocean_usd",
                "expense_category": "中国海运费",
                "allocation_basis": "gross_weight",
                "currency": "USD",
                "amount": 30,
            }
        ],
        items=[
            {
                "row_no": 1,
                "material_code": "YL000001",
                "product_name": "太阳眼镜",
                "quantity": 2,
                "actual_shipped_qty": 2,
                "unit_price": 8,
                "purchase_currency": "RMB",
                "goods_value": 16,
                "freight_alloc_rmb": 6,
                "import_tax_total": 3,
                "derived_json": json.dumps({"fx_rmb_to_mxn": 3}),
                "total_cost_rmb": 25,
                "total_unit_rmb": 12.5,
            }
        ],
    )

    assert payload["target_system"] == "DeepLinkERP"
    assert payload["subsidiary_code"] == "MX01"
    assert payload["items"][0]["material_code"] == "YL000001"
    assert payload["items"][0]["original_unit_price"] == 8
    assert payload["items"][0]["comprehensive_unit_price"] == 12.5
    assert payload["items"][0]["outbound_quantity"] == 2
    assert payload["items"][0]["expense_detail"]["logistics"]["freight_alloc_rmb"] == 6
    clearance_tax = payload["items"][0]["expense_detail"]["clearance_and_tax"]
    assert clearance_tax["clearance_alloc_rmb"] == 2
    assert clearance_tax["tax_alloc_rmb"] == 1


def test_writeback_to_erp_records_failed_attempt_when_config_missing(monkeypatch) -> None:
    from overseas_costing.services import batch_service

    set_values = []
    inserted_logs = []

    class FakeDB:
        @staticmethod
        def set_value(doctype, name, values, update_modified=False):
            set_values.append((doctype, name, values, update_modified))

        @staticmethod
        def commit():
            pass

    class FakeDoc:
        def __init__(self, payload):
            self.payload = payload

        def insert(self, ignore_permissions=False):
            inserted_logs.append((self.payload, ignore_permissions))
            return self

    class FakeSession:
        user = "tester@example.com"

    class FakeFrappe:
        db = FakeDB()
        session = FakeSession()

        @staticmethod
        def get_doc(payload):
            return FakeDoc(payload)

    monkeypatch.setattr(batch_service, "frappe", FakeFrappe)
    monkeypatch.setattr(
        batch_service,
        "preview_erp_payload",
        lambda batch_name, version_name=None: {
            "ok": True,
            "ready": True,
            "batch_name": "BATCH-001",
            "version_name": "VERSION-001",
            "payload": {"target_system": "DeepLinkERP", "batch_no": "BATCH-001", "item_count": 1, "subsidiary_code": "MX01"},
        },
    )
    monkeypatch.setattr(
        batch_service,
        "_load_erp_push_context",
        lambda batch_name, version_name=None: {"batch": {"extra_json": "{}", "confirm_status": "Confirmed"}},
    )
    monkeypatch.setattr(
        batch_service.erp_client,
        "push_overseas_cost_payload",
        lambda payload: {
            "ok": False,
            "status": "Failed",
            "message": "缺少 DeepLinkERP 目标 DocType 配置",
            "request": {"authorization_configured": True},
            "response": {},
        },
    )

    result = writeback_to_erp("BATCH-001", "VERSION-001")

    assert result["ok"] is False
    assert result["writeback_status"] == "Failed"
    assert result["retryable"] is True
    assert set_values[0][2]["writeback_status"] == "Failed"
    assert set_values[0][2]["writeback_message"] == "缺少 DeepLinkERP 目标 DocType 配置"
    saved_extra = json.loads(set_values[0][2]["extra_json"])
    assert saved_extra["erp_writeback"]["attempt_count"] == 1
    assert saved_extra["erp_writeback"]["attempt_history"][0]["status"] == "Failed"
    assert inserted_logs[0][0]["action_type"] == "WRITEBACK"


def test_writeback_to_erp_blocks_when_not_confirmed(monkeypatch) -> None:
    from overseas_costing.services import batch_service

    set_values = []

    class FakeDB:
        @staticmethod
        def set_value(doctype, name, values, update_modified=False):
            set_values.append((doctype, name, values, update_modified))

        @staticmethod
        def commit():
            pass

    class FakeSession:
        user = "tester@example.com"

    class FakeFrappe:
        db = FakeDB()
        session = FakeSession()

    monkeypatch.setattr(batch_service, "frappe", FakeFrappe)
    monkeypatch.setattr(
        batch_service,
        "preview_erp_payload",
        lambda batch_name, version_name=None: {
            "ok": True,
            "ready": True,
            "batch_name": "BATCH-001",
            "version_name": "VERSION-001",
            "payload": {"target_system": "DeepLinkERP", "batch_no": "BATCH-001", "item_count": 1, "subsidiary_code": "MX01"},
        },
    )
    monkeypatch.setattr(
        batch_service,
        "_load_erp_push_context",
        lambda batch_name, version_name=None: {"batch": {"extra_json": "{}", "confirm_status": "Pending", "writeback_status": "Not Started"}},
    )
    monkeypatch.setattr(
        batch_service.erp_client,
        "push_overseas_cost_payload",
        lambda payload: (_ for _ in ()).throw(AssertionError("push_overseas_cost_payload should not run")),
    )

    result = writeback_to_erp("BATCH-001", "VERSION-001")

    assert result["ok"] is False
    assert result["queued"] is False
    assert result["pushed"] is False
    assert result["retryable"] is False
    assert "校验计算结果" in result["message"]
    assert set_values == []


def test_writeback_to_erp_records_success_and_target_doc(monkeypatch) -> None:
    from overseas_costing.services import batch_service

    set_values = []

    class FakeDB:
        @staticmethod
        def set_value(doctype, name, values, update_modified=False):
            set_values.append((doctype, name, values, update_modified))

        @staticmethod
        def commit():
            pass

    class FakeDoc:
        def __init__(self, payload):
            self.payload = payload

        def insert(self, ignore_permissions=False):
            return self

    class FakeSession:
        user = "tester@example.com"

    class FakeFrappe:
        db = FakeDB()
        session = FakeSession()

        @staticmethod
        def get_doc(payload):
            return FakeDoc(payload)

    previous_extra = {
        "erp_writeback": {
            "attempt_count": 1,
            "attempt_history": [{"attempt_no": 1, "status": "Failed"}],
        }
    }
    monkeypatch.setattr(batch_service, "frappe", FakeFrappe)
    monkeypatch.setattr(
        batch_service,
        "preview_erp_payload",
        lambda batch_name, version_name=None: {
            "ok": True,
            "ready": True,
            "batch_name": "BATCH-001",
            "version_name": "VERSION-001",
            "payload": {"target_system": "DeepLinkERP", "batch_no": "BATCH-001", "item_count": 1, "subsidiary_code": "MX01"},
        },
    )
    monkeypatch.setattr(
        batch_service,
        "_load_erp_push_context",
        lambda batch_name, version_name=None: {
            "batch": {
                "extra_json": json.dumps(previous_extra, ensure_ascii=False),
                "confirm_status": "Confirmed",
            }
        },
    )
    monkeypatch.setattr(
        batch_service.erp_client,
        "push_overseas_cost_payload",
        lambda payload: {
            "ok": True,
            "status": "Success",
            "message": "DeepLinkERP 返回成功，目标单据 ERP-PUSH-001。",
            "erp_target_doc": "ERP-PUSH-001",
            "http_status": 200,
            "request": {"target_doctype": "Overseas Cost Push"},
            "response": {"data": {"name": "ERP-PUSH-001"}},
        },
    )

    result = writeback_to_erp("BATCH-001", "VERSION-001")

    assert result["ok"] is True
    assert result["pushed"] is True
    assert result["writeback_status"] == "Success"
    assert result["erp_target_doc"] == "ERP-PUSH-001"
    assert set_values[0][2]["writeback_status"] == "Success"
    assert set_values[0][2]["erp_target_doc"] == "ERP-PUSH-001"
    saved_extra = json.loads(set_values[0][2]["extra_json"])
    assert saved_extra["erp_writeback"]["attempt_count"] == 2
    assert [item["status"] for item in saved_extra["erp_writeback"]["attempt_history"]] == ["Failed", "Success"]


def test_build_writeback_field_gaps_exposes_missing_item_fieldnames() -> None:
    gaps = _build_writeback_field_gaps(
        batch={
            "subsidiary_code": "",
            "current_version": "",
            "estimated_total_cost_rmb": 0,
            "actual_total_cost_rmb": 0,
        },
        items=[
            {
                "name": "ITEM-001",
                "row_no": 1,
                "material_code": "",
                "product_name": "澶槼鐪奸暅",
                "quantity": 0,
                "actual_shipped_qty": 0,
                "unit_price": 0,
                "purchase_currency": "",
                "goods_value": 0,
                "total_unit_rmb": 0,
            }
        ],
        rules=[],
        resolved_version_name=None,
    )

    assert gaps["items"][0]["missing_fieldnames"]


def test_hidden_approval_status_matches_revoked_dingtalk_statuses() -> None:
    assert is_hidden_approval_status("TERMINATED") is True
    assert is_hidden_approval_status("已撤销") is True
    assert is_hidden_approval_status("COMPLETED") is False
    assert is_hidden_approval_status("") is False
    assert is_invalid_approval_status("REJECTED") is True
    assert is_invalid_approval_status("已驳回") is True
    assert is_invalid_approval_status("COMPLETED") is False


def test_build_batch_source_status_summarizes_oa_and_voucher_records() -> None:
    status = _build_batch_source_status(
        {
            "name": "BATCH-001",
            "batch_no": "FSCU8486789",
            "source_type": "oa_logistics",
            "source_approval_no": "202607010001",
            "source_attachment_count": 3,
        },
        [
            {"source_type": "OA", "attachment_type": "Packing List", "parse_status": "Queued"},
            {"source_type": "OA", "attachment_type": "Commercial Invoice", "parse_status": "Queued"},
            {
                "name": "ATT-OLD",
                "source_type": "Voucher",
                "attachment_type": "Tax Certificate",
                "parse_status": "Parsed",
                "modified": "2026-07-01 10:00:00",
                "mapped_result_json": json.dumps(
                    {
                        "status": "passed",
                        "status_label": "一致",
                        "voucher": {"paid_total_mxn": 100},
                        "system": {"system_import_tax_total_mxn": 100},
                        "difference": {"tax_total_diff_mxn": 0, "direction_label": "一致"},
                    },
                    ensure_ascii=False,
                ),
            },
            {
                "name": "ATT-NEW",
                "file_name": "PD_MZ260108凭证.pdf",
                "source_type": "Voucher",
                "attachment_type": "Tax Certificate",
                "parse_status": "Parsed",
                "modified": "2026-07-20 10:00:00",
                "parse_result_json": json.dumps({"summary": {"paid_total_mxn": 129883}}, ensure_ascii=False),
                "mapped_result_json": json.dumps(
                    {
                        "status": "review",
                        "status_label": "需复核",
                        "system": {"system_import_tax_total_mxn": 130186},
                        "difference": {"tax_total_diff_mxn": -303, "direction_label": "凭证金额低于系统"},
                    },
                    ensure_ascii=False,
                ),
            },
        ],
    )

    assert status["has_oa_logistics"] is True
    assert status["source_no"] == "202607010001"
    assert status["oa_attachment_count"] == 3
    assert status["registered_attachment_count"] == 4
    assert status["packing_list_count"] == 1
    assert status["parsed_packing_list_count"] == 0
    assert status["tax_certificate_count"] == 2
    assert status["parsed_tax_certificate_count"] == 2
    assert status["latest_tax_certificate_reconciliation"] == {
        "name": "ATT-NEW",
        "file_name": "PD_MZ260108凭证.pdf",
        "modified": "2026-07-20 10:00:00",
        "status": "review",
        "status_label": "需复核",
        "manual_resolution_status_label": "",
        "paid_total_mxn": 129883,
        "system_tax_total_mxn": 130186,
        "tax_total_diff_mxn": -303,
        "direction_label": "凭证金额低于系统",
    }


def test_build_batch_source_status_exposes_quote_candidates_without_raw_oa_text() -> None:
    status = _build_batch_source_status(
        {
            "name": "BATCH-QUOTE-001",
            "extra_json": json.dumps(
                {
                    "source": "dingtalk_oa_logistics",
                    "logistics_quote_candidates": [
                        {
                            "carrier": "SISA",
                            "amount": 5730,
                            "currency": "RMB",
                            "volume_m3": 1.5,
                            "evidence_line": "合计价格：5730元",
                            "source_field": "物流报价",
                            "source_value": "不应暴露给前端的完整原文",
                        }
                    ],
                    "confirmed_logistics_quote": {
                        "carrier": "SISA",
                        "amount": 5730,
                        "currency": "RMB",
                    },
                },
                ensure_ascii=False,
            ),
        }
    )

    assert status["logistics_quote_candidate_count"] == 1
    quote = status["logistics_quote_candidates"][0]
    assert quote["carrier"] == "SISA"
    assert quote["amount"] == 5730
    assert quote["currency"] == "RMB"
    assert quote["volume_m3"] == 1.5
    assert quote["evidence_line"] == "合计价格：5730元"
    assert quote["source_field"] == "物流报价"
    assert quote["status"] == "待确认"
    assert status["has_confirmed_logistics_quote"] is True
    assert "source_value" not in quote


def test_build_batch_source_status_exposes_logistics_text_summary() -> None:
    status = _build_batch_source_status(
        {
            "name": "BATCH-DHL-001",
            "source_type": "oa_logistics",
            "extra_json": json.dumps(
                {
                    "source": "dingtalk_oa_logistics",
                    "form_fields": {
                        "物流报价Cotización de logística": "DHL报价：\n50.22*1.3825*43.2+155*1.3825*1+155*1=3368.62678元",
                        "物流方式Camino Envío": "Express快递",
                        "预计发货日期Fecha de Pre-entrega": "2026/8/12",
                        "目标地区Países destinatarios": "MANZANILLO Mexico",
                        "重量Peso（KG）": "43.2",
                    },
                },
                ensure_ascii=False,
            ),
        }
    )

    assert status["logistics_text_summary"]["transport_mode"] == "EXPRESS"
    assert status["logistics_text_summary"]["logistics_quote_carrier"] == "DHL"
    assert status["logistics_text_summary"]["logistics_quote_amount"] == 3368.62678
    assert status["logistics_text_summary"]["pre_delivery_date"] == "2026/8/12"
    assert status["logistics_text_summary"]["destination"] == "MANZANILLO Mexico"
    assert "logistics_quote_text" not in status["logistics_text_summary"]

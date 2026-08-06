"""中文用途：调试脚本测试。"""

from pathlib import Path

from openpyxl import load_workbook

from overseas_costing.scripts import compare_manual_excel_baseline
from overseas_costing.scripts import import_excel_workbook
from overseas_costing.scripts.recalculate_batch import recalculate, recalculate_from_env


def test_recalculate_script_delegates_to_service_in_dry_run() -> None:
    result = recalculate(batch_name="BATCH-001", version_name="VERSION-001")

    assert result["ok"] is True
    assert result["dry_run"] is True
    assert result["batch_name"] == "BATCH-001"
    assert result["version_name"] == "VERSION-001"
    assert "summary_snapshot" in result


def test_recalculate_script_reads_batch_from_env(monkeypatch) -> None:
    monkeypatch.setenv("OVERSEAS_COST_BATCH", "BATCH-002")
    monkeypatch.setenv("OVERSEAS_COST_VERSION", "VERSION-002")

    result = recalculate_from_env()

    assert result["ok"] is True
    assert result["batch_name"] == "BATCH-002"
    assert result["version_name"] == "VERSION-002"


def test_recalculate_script_requires_batch_env(monkeypatch) -> None:
    monkeypatch.delenv("OVERSEAS_COST_BATCH", raising=False)

    result = recalculate_from_env()

    assert result["ok"] is False
    assert "OVERSEAS_COST_BATCH" in result["message"]


def test_import_excel_workbook_preview_from_env(monkeypatch) -> None:
    workbook_path = Path("attachment.xlsx")

    def fake_parse(path: Path, sheet_name: str | None = None):
        assert Path(path) == workbook_path
        assert sheet_name == "7月份钢化膜空运"
        return (
            {"sourceSheet": "7月份钢化膜空运", "parser": "oa_attachment_detail"},
            [
                {
                    "id": "PO-001",
                    "batchNo": "PO-001",
                    "sourceSheet": "7月份钢化膜空运",
                    "transportMode": "空运",
                    "items": [["FL004106", "钢化膜", 1.2, 500, 600]],
                }
            ],
        )

    monkeypatch.setattr(import_excel_workbook, "parse_yuewei_excel_workbook", fake_parse)
    monkeypatch.setattr(import_excel_workbook, "_resolve_file_path", lambda file_path=None: workbook_path)
    monkeypatch.setenv("OVERSEAS_COST_EXCEL_FILE", str(workbook_path))
    monkeypatch.setenv("OVERSEAS_COST_SOURCE_SHEET", "7月份钢化膜空运")
    monkeypatch.setenv("OVERSEAS_COST_TRANSPORT_KEYWORD", "空运")

    result = import_excel_workbook.preview_from_env()

    assert result["ok"] is True
    assert result["parser_meta"]["parser"] == "oa_attachment_detail"
    assert result["selected_summary"]["block_count"] == 1
    assert result["selected_summary"]["item_count"] == 1
    assert result["selection"]["include_double_clear"] is True


def test_import_excel_workbook_import_from_env(monkeypatch) -> None:
    workbook_path = Path("cost.xlsx")

    def fake_import(**kwargs):
        return {
            "ok": True,
            "source_name": kwargs["source_name"],
            "file_path": kwargs["file_path"],
            "source_sheet": kwargs["source_sheet"],
            "transport_keyword": kwargs["transport_keyword"],
            "limit": kwargs["limit"],
        }

    monkeypatch.setattr(import_excel_workbook, "import_yuewei_excel_file", fake_import)
    monkeypatch.setattr(import_excel_workbook, "_resolve_file_path", lambda file_path=None: workbook_path)
    monkeypatch.setenv("OVERSEAS_COST_EXCEL_FILE", str(workbook_path))
    monkeypatch.setenv("OVERSEAS_COST_SOURCE_NAME", "7月份钢化膜空运明细.xlsx")
    monkeypatch.setenv("OVERSEAS_COST_SOURCE_SHEET", "7月份钢化膜空运")
    monkeypatch.setenv("OVERSEAS_COST_TRANSPORT_KEYWORD", "空运")
    monkeypatch.setenv("OVERSEAS_COST_LIMIT", "2")

    result = import_excel_workbook.import_from_env()

    assert result["ok"] is True
    assert result["source_name"] == "7月份钢化膜空运明细.xlsx"
    assert result["source_sheet"] == "7月份钢化膜空运"
    assert result["transport_keyword"] == "空运"
    assert result["limit"] == 2


def test_compare_manual_excel_baseline_matches_and_summarizes() -> None:
    manual_rows = [
        {
            "manual_excel_row": 79,
            "material_code": "FL001",
            "product_name": "太阳眼镜",
            "quantity": 2,
            "goods_value": 100,
            "gross_weight_kg": 5,
            "freight_alloc_rmb": 10,
            "mexico_customs_rmb": 20,
            "total_cost_rmb": 130,
            "total_unit_rmb": 65,
        },
        {
            "manual_excel_row": 80,
            "material_code": "FL002",
            "product_name": "包装盒",
            "quantity": 1,
            "goods_value": 50,
            "gross_weight_kg": 3,
            "freight_alloc_rmb": 5,
            "mexico_customs_rmb": 10,
            "total_cost_rmb": 65,
            "total_unit_rmb": 65,
        },
    ]
    system_rows = [
        {
            "row_no": 1,
            "material_code": "FL001",
            "product_name": "太阳眼镜",
            "quantity": 2,
            "goods_value": 100,
            "gross_weight_kg": 5,
            "freight_alloc_rmb": 10,
            "mexico_customs_rmb": 20,
            "total_cost_rmb": 130,
            "total_unit_rmb": 65,
        },
        {
            "row_no": 2,
            "material_code": "FL003",
            "product_name": "未匹配物料",
            "quantity": 1,
            "goods_value": 40,
            "gross_weight_kg": 2,
            "freight_alloc_rmb": 4,
            "mexico_customs_rmb": 8,
            "total_cost_rmb": 52,
            "total_unit_rmb": 52,
        },
    ]

    comparison = compare_manual_excel_baseline.build_comparison_rows(manual_rows, system_rows)
    summary = compare_manual_excel_baseline.build_summary_rows(manual_rows, system_rows)

    assert comparison["matched_count"] == 1
    assert comparison["manual_unmatched_count"] == 1
    assert len(comparison["unmatched_system_rows"]) == 1
    assert comparison["rows"][0]["match_status"] == "已匹配"
    assert comparison["rows"][1]["match_status"] == "人工表有，系统未匹配"
    assert summary[0] == {
        "field": "item_count",
        "label": "物料行数",
        "manual": 2,
        "system": 2,
        "diff": 0.0,
        "diff_pct": 0.0,
        "status": "一致",
    }


def test_compare_manual_excel_baseline_matches_same_code_by_spec_model() -> None:
    manual_rows = [
        {
            "manual_excel_row": 79,
            "material_code": "YL000098",
            "product_name": "TPU-HF-8695AU",
            "quantity": 5000,
            "goods_value": 71787.5,
            "total_cost_rmb": 124471.0,
        },
        {
            "manual_excel_row": 80,
            "material_code": "YL000098",
            "product_name": "TPU-HF-1190A-1",
            "quantity": 8000,
            "goods_value": 107968.4,
            "total_cost_rmb": 129838.4,
        },
    ]
    system_rows = [
        {
            "row_no": 1,
            "material_code": "YL000098",
            "product_name": "TPU原料",
            "spec_model": "HF-8695AU",
            "quantity": 200,
            "goods_value": 13500,
            "total_cost_rmb": 20598.1,
        },
        {
            "row_no": 2,
            "material_code": "YL000098",
            "product_name": "TPU原料",
            "spec_model": "HF-1190A-1",
            "quantity": 320,
            "goods_value": 14800,
            "total_cost_rmb": 24078.2,
        },
    ]

    comparison = compare_manual_excel_baseline.build_comparison_rows(manual_rows, system_rows)

    assert comparison["matched_count"] == 2
    assert comparison["manual_unmatched_count"] == 0
    assert len(comparison["unmatched_system_rows"]) == 0
    assert comparison["rows"][0]["system_spec_model"] == "HF-8695AU"
    assert "按物料编码+规格匹配" in comparison["rows"][0]["remark"]


def test_compare_manual_excel_baseline_exports_workbook(monkeypatch) -> None:
    manual_rows = [
        {
            "manual_excel_row": 79,
            "material_code": "FL001",
            "product_name": "太阳眼镜",
            "quantity": 2,
            "goods_value": 100,
            "gross_weight_kg": 5,
            "freight_alloc_rmb": 10,
            "mexico_customs_rmb": 20,
            "total_cost_rmb": 130,
            "total_unit_rmb": 65,
        }
    ]
    output_path = Path("data") / "_test_compare_manual_excel_baseline.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.unlink(missing_ok=True)

    monkeypatch.setattr(compare_manual_excel_baseline, "_resolve_manual_workbook", lambda _file_path=None: Path("manual.xlsx"))
    monkeypatch.setattr(
        compare_manual_excel_baseline,
        "load_manual_baseline",
        lambda **_kwargs: {
            "source_file": "manual.xlsx",
            "source_sheet": "2026年YUEWEI",
            "source_range": "2026年YUEWEI!79:79",
            "customs_no": "26 16 1681 6000151",
            "waybill_no": "HPCU5155607",
            "items": manual_rows,
        },
    )
    monkeypatch.setattr(
        compare_manual_excel_baseline.batch_service,
        "get_batch_items",
        lambda **_kwargs: {
            "ok": True,
            "batch_name": "HPCU5155607",
            "version_name": "VER-001",
            "items": [
                {
                    "row_no": 1,
                    "material_code": "FL001",
                    "product_name": "太阳眼镜",
                    "quantity": 2,
                    "goods_value": 100,
                    "gross_weight_kg": 5,
                    "freight_alloc_rmb": 10,
                    "mexico_customs_rmb": 20,
                    "total_cost_rmb": 130,
                    "total_unit_rmb": 65,
                }
            ],
        },
    )

    result = compare_manual_excel_baseline.build_hpcu_manual_comparison(output_path=str(output_path))
    try:
        assert result["ok"] is True
        assert result["manual_item_count"] == 1
        assert result["system_item_count"] == 1
        assert result["matched_count"] == 1
        workbook = load_workbook(output_path, data_only=True)
        assert workbook.sheetnames == ["汇总差异", "逐行对照", "系统未匹配明细"]
        assert workbook["汇总差异"]["A1"].value == "项目"
        assert workbook["逐行对照"]["A2"].value == "已匹配"
    finally:
        if "workbook" in locals():
            workbook.close()
        output_path.unlink(missing_ok=True)

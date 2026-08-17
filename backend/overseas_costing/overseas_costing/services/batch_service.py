"""
中文用途：批次查询与回写检查服务。

当前先返回稳定的数据结构，占住后续前端对接形态。
后面接数据库时，优先从这里补查询实现。
"""

from __future__ import annotations

import base64
from datetime import datetime, timedelta
from io import BytesIO
import json

from overseas_costing.services import source_priority_service

try:
    import frappe
except Exception:  # pragma: no cover - 本地无 Frappe 环境时保持可导入
    frappe = None

from overseas_costing.utils.dingtalk import build_dingtalk_order_payload, extract_dingtalk_instance_id


DEFAULT_BATCH_RECENT_DAYS = 30
CLASSIC_HISTORY_BATCH_KEYS = (
    "HPCU5155607",
    "FSCU8486789",
    "MXT959831",
)


def _get_batch_source_meta(batch_name: str) -> dict:
    """读取批次上的钉钉原单跳转元信息。"""

    meta = {
        "name": batch_name,
        "batch_no": "",
        "source_approval_no": "",
        "source_instance_id": "",
        "source_dingtalk_url": "",
        "source_title": "",
        "source_creator_name": "",
        "source_approval_status": "",
    }
    if frappe is None:
        return meta

    fields = list(meta.keys())
    try:
        rows = frappe.get_all(
            "Overseas Cost Batch",
            filters={"name": batch_name},
            fields=fields,
            limit_page_length=1,
        )
        if not rows:
            rows = frappe.get_all(
                "Overseas Cost Batch",
                filters={"batch_no": batch_name},
                fields=fields,
                limit_page_length=1,
            )
    except Exception:
        return meta

    if rows:
        meta.update(rows[0])
        if not meta.get("source_instance_id") and not meta.get("source_dingtalk_url"):
            try:
                item_rows = frappe.get_all(
                    "Overseas Cost Item",
                    filters={"batch": meta.get("name")},
                    fields=["source_doc_no", "dingtalk_instance_id", "dingtalk_official_url"],
                    limit_page_length=50,
                )
            except Exception:
                item_rows = []
            for item in item_rows:
                if item.get("dingtalk_instance_id") or item.get("dingtalk_official_url"):
                    meta["source_approval_no"] = meta.get("source_approval_no") or item.get("source_doc_no") or ""
                    meta["source_instance_id"] = item.get("dingtalk_instance_id") or extract_dingtalk_instance_id(item.get("dingtalk_official_url")) or ""
                    meta["source_dingtalk_url"] = item.get("dingtalk_official_url") or ""
                    break
    if not meta.get("source_instance_id"):
        meta["source_instance_id"] = extract_dingtalk_instance_id(meta.get("source_dingtalk_url"))
    return meta


def get_batch_list(filters: dict) -> dict:
    return {
        "ok": True,
        "message": "批次列表接口骨架已创建，后续接数据库查询。",
        "filters": filters,
        "items": [],
        "total": 0,
    }


def get_batch_detail(batch_name: str, version_name: str | None = None) -> dict:
    source_meta = _get_batch_source_meta(batch_name)
    return {
        "ok": True,
        "message": "批次详情接口骨架已创建。",
        "batch_name": batch_name,
        "version_name": version_name,
        "header": {
            "batch_no": source_meta.get("batch_no") or batch_name,
            "source_approval_no": source_meta.get("source_approval_no", ""),
            "source_instance_id": source_meta.get("source_instance_id", ""),
            "source_dingtalk_url": source_meta.get("source_dingtalk_url", ""),
            "source_title": source_meta.get("source_title", ""),
            "source_creator_name": source_meta.get("source_creator_name", ""),
            "source_approval_status": source_meta.get("source_approval_status", ""),
        },
        "summary": {},
    }


def get_batch_items(batch_name: str, version_name: str | None = None) -> dict:
    return {
        "ok": True,
        "message": "批次明细接口骨架已创建。",
        "batch_name": batch_name,
        "version_name": version_name,
        "columns": [],
        "items": [],
    }


def get_version_list(batch_name: str) -> dict:
    return {
        "ok": True,
        "message": "版本列表接口骨架已创建。",
        "batch_name": batch_name,
        "items": [],
    }


def get_dingtalk_order_link(batch_name: str) -> dict:
    """返回钉钉原单跳转信息。"""

    source_meta = _get_batch_source_meta(batch_name)
    payload = build_dingtalk_order_payload(
        batch_name=source_meta.get("batch_no") or batch_name,
        approval_no=source_meta.get("source_approval_no"),
        instance_id=source_meta.get("source_instance_id"),
        official_url=source_meta.get("source_dingtalk_url"),
    )
    return {
        "ok": True,
        "batch_name": batch_name,
        "message": "钉钉原单跳转信息已生成。" if payload["can_open"] else "当前批次缺少钉钉实例ID或官方链接。",
        "dingtalk_order": payload,
    }


def check_writeback_ready(batch_name: str, version_name: str | None = None) -> dict:
    return {
        "ok": True,
        "ready": False,
        "batch_name": batch_name,
        "version_name": version_name,
        "checks": {
            "has_current_version": False,
            "is_confirmed": False,
            "has_dirty_data": True,
        },
        "message": "回写检查骨架已创建，后续接正式校验规则。",
    }


def writeback_to_erp(batch_name: str, version_name: str) -> dict:
    return {
        "ok": True,
        "queued": True,
        "batch_name": batch_name,
        "version_name": version_name,
        "message": "ERP 回写骨架已创建，当前仅预留入口。",
    }


# --- Database-backed query implementation for the Excel flat-table MVP. ---

import json as _json

from overseas_costing.utils.field_mapper import normalize_transport_mode

EXCEL_COLUMNS = [
    {"excel_col": "A", "fieldname": "material_code", "label": "物料编码"},
    {"excel_col": "B", "fieldname": "product_name", "label": "产品名称"},
    {"excel_col": "B1", "fieldname": "spec_model", "label": "规格型号 Especificación / Modelo"},
    {"excel_col": "C", "fieldname": "unit_price", "label": "采购单价"},
    {"excel_col": "C1", "fieldname": "purchase_currency", "label": "采购币种"},
    {"excel_col": "D", "fieldname": "quantity", "label": "采购数量"},
    {"excel_col": "D1", "fieldname": "unit", "label": "单位"},
    {"excel_col": "E", "fieldname": "goods_value", "label": "总货值"},
    {"excel_col": "F", "fieldname": "import_name", "label": "海关进口名称"},
    {"excel_col": "G", "fieldname": "hs_code", "label": "海关分类编码"},
    {"excel_col": "H", "fieldname": "category", "label": "大类分类"},
    {"excel_col": "I", "fieldname": "customs_no", "label": "报关单号"},
    {"excel_col": "J", "fieldname": "waybill_no", "label": "中国到墨西哥运单号"},
    {"excel_col": "K", "fieldname": "china_misc_rmb", "label": "中国运输及相关杂费 RMB"},
    {"excel_col": "L", "fieldname": "china_misc_mxn", "label": "中国运输及相关杂费 MXN"},
    {"excel_col": "M", "fieldname": "china_ocean_usd", "label": "中国海运 USD"},
    {"excel_col": "N", "fieldname": "cc_rate", "label": "C.C税率"},
    {"excel_col": "O", "fieldname": "cc_anti_dumping", "label": "C.C反倾销附加税"},
    {"excel_col": "P", "fieldname": "igi_rate", "label": "IGI关税税率"},
    {"excel_col": "Q", "fieldname": "igi_amount", "label": "IGI关税税额"},
    {"excel_col": "R", "fieldname": "iva_rate", "label": "IVA增值税税率"},
    {"excel_col": "S", "fieldname": "iva_amount", "label": "IVA增值税税额"},
    {"excel_col": "T", "fieldname": "goods_value_ratio", "label": "分摊比例（货值比）"},
    {"excel_col": "U", "fieldname": "dta", "label": "DTA海关递延税款"},
    {"excel_col": "V", "fieldname": "prv_duty", "label": "PRV从价税/关税"},
    {"excel_col": "W", "fieldname": "prv_iva", "label": "PRV的IVA"},
    {"excel_col": "X", "fieldname": "import_tax_total", "label": "IMPUESTOS合计清关税费"},
    {"excel_col": "Y", "fieldname": "revalidacion", "label": "REVALIDACION文件验证费"},
    {"excel_col": "Z", "fieldname": "maniobras", "label": "MANIOBRAS码头操作费"},
    {"excel_col": "AA", "fieldname": "muellaje", "label": "MUELLAJE码头作业费"},
    {"excel_col": "AB", "fieldname": "entrega_mercancia", "label": "ENTREGA DE MERCANCIA配送费用"},
    {"excel_col": "AC", "fieldname": "previo", "label": "PREVIO预检费用"},
    {"excel_col": "AD", "fieldname": "service_aa", "label": "Servicios A.A.货代服务费"},
    {"excel_col": "AE", "fieldname": "almacenajes", "label": "ALMACENAJES仓储费"},
    {"excel_col": "AF", "fieldname": "reconocimiento_aduanero", "label": "RECONOCIMIENTO ADUANERO"},
    {"excel_col": "AG", "fieldname": "honorarios", "label": "HONORARIOS"},
    {"excel_col": "AH", "fieldname": "complemento_maniobras", "label": "COMPLEMENTO DE MANIOBRAS"},
    {"excel_col": "AI", "fieldname": "desconsolidacion", "label": "DESCONSOLIDACION"},
    {"excel_col": "AJ", "fieldname": "maniobra_falso", "label": "MANIOBRA EN FALSO"},
    {"excel_col": "AK", "fieldname": "arrastre", "label": "ARRASTRE拖运费"},
    {"excel_col": "AL", "fieldname": "patio_regulador", "label": "PATIO REGULADOR"},
    {"excel_col": "AM", "fieldname": "entrega_vacio", "label": "ENTREGA DE VACIO还箱费"},
    {"excel_col": "AN", "fieldname": "limpieza_contenedor", "label": "LIMPIEZA DE CONTENEDOR"},
    {"excel_col": "AO", "fieldname": "mexico_customs_mxn", "label": "墨西哥清关费用 MXN"},
    {"excel_col": "AP", "fieldname": "mexico_customs_rmb", "label": "墨西哥清关费用 RMB"},
    {"excel_col": "AQ", "fieldname": "mexico_customs_usd", "label": "墨西哥清关费用 USD"},
    {"excel_col": "AR", "fieldname": "mexico_inland_mxn", "label": "墨西哥内陆运输费用 MXN"},
    {"excel_col": "AS", "fieldname": "mexico_misc_mxn", "label": "墨西哥杂费 MXN"},
    {"excel_col": "AT", "fieldname": "mexico_inland_misc_rmb", "label": "墨西哥内陆运输+杂费 RMB"},
    {"excel_col": "AU", "fieldname": "china_to_mexico_freight_rmb", "label": "中国到墨西哥运费 RMB"},
    {"excel_col": "AV", "fieldname": "gross_weight_kg", "label": "货重毛重 KG"},
    {"excel_col": "AW", "fieldname": "weight_ratio", "label": "分摊比例（重量比）"},
    {"excel_col": "AX", "fieldname": "freight_alloc_rmb", "label": "运输费用分摊 RMB"},
    {"excel_col": "AY", "fieldname": "freight_alloc_mxn", "label": "运输费用分摊 MXN"},
    {"excel_col": "AZ", "fieldname": "total_logistics_mxn", "label": "运输+清关+杂费 MXN"},
    {"excel_col": "BA", "fieldname": "alloc_price_mxn", "label": "分摊物流价格 MXN"},
    {"excel_col": "BB", "fieldname": "total_cost_rmb", "label": "综合成本 RMB"},
    {"excel_col": "BC", "fieldname": "total_unit_rmb", "label": "综合单价 RMB"},
    {"excel_col": "BD", "fieldname": "project_collection", "label": "项目归集"},
    {"excel_col": "BE", "fieldname": "transport_mode", "label": "运输方式"},
]
EXCEL_FIELDNAMES = [column["fieldname"] for column in EXCEL_COLUMNS]
EXTRA_ITEM_FIELDS = [
    "name",
    "row_no",
    "excel_row_no",
    "product_name_es",
    "spec_model",
    "unit",
    "recipient",
    "purchase_currency",
    "actual_shipped_qty",
    "volume_m3",
    "volume_weight_kg",
    "chargeable_weight_kg",
    "source_type",
    "source_doc_no",
    "source_file_name",
    "parse_status",
    "manual_override_flag",
    "dingtalk_instance_id",
    "dingtalk_official_url",
    "derived_json",
]
ITEM_FILTER_FIELDS = (
    "customs_no",
    "waybill_no",
    "material_code",
    "product_name",
    "import_name",
    "hs_code",
    "category",
)
ITEM_KEYWORD_FIELDS = ITEM_FILTER_FIELDS + ("project_collection", "transport_mode")
DEFAULT_FX_RMB_TO_MXN = 2.6
DEFAULT_FX_USD_TO_RMB = round(1 / 0.1393, 6)
HIDDEN_APPROVAL_STATUSES = ("TERMINATED", "CANCELED", "CANCELLED", "REVOKED", "撤销", "已撤销")


def is_hidden_approval_status(status: str | None) -> bool:
    """判断钉钉审批状态是否不应在成本表格中展示。"""

    normalized = str(status or "").strip().upper()
    if not normalized:
        return False
    return any(str(hidden).upper() in normalized for hidden in HIDDEN_APPROVAL_STATUSES)


def _load_batch_payload(batch_payload: str | dict | None) -> dict:
    if not batch_payload:
        return {}
    if isinstance(batch_payload, dict):
        return dict(batch_payload)
    loaded = _json.loads(batch_payload)
    if not isinstance(loaded, dict):
        raise ValueError("新增报关运单参数必须是对象。")
    return loaded


def _clean_payload_text(payload: dict, fieldname: str) -> str:
    return str(payload.get(fieldname) or "").strip()


def _build_manual_batch_values(payload: dict) -> dict:
    batch_no = _clean_payload_text(payload, "batch_no")
    source_dingtalk_url = _clean_payload_text(payload, "source_dingtalk_url")
    source_instance_id = _clean_payload_text(payload, "source_instance_id") or extract_dingtalk_instance_id(source_dingtalk_url)
    values = {
        "batch_no": batch_no,
        "customs_no": _clean_payload_text(payload, "customs_no"),
        "waybill_no": _clean_payload_text(payload, "waybill_no"),
        "container_no": _clean_payload_text(payload, "container_no"),
        "sea_bill_no": _clean_payload_text(payload, "sea_bill_no"),
        "commercial_invoice_no": _clean_payload_text(payload, "commercial_invoice_no"),
        "transport_mode": normalize_transport_mode(payload.get("transport_mode")) or "SEA",
        "project_collection": _clean_payload_text(payload, "project_collection"),
        "source_type": "manual",
        "source_approval_no": _clean_payload_text(payload, "source_approval_no"),
        "source_instance_id": source_instance_id,
        "source_dingtalk_url": source_dingtalk_url,
        "status": "Draft",
        "confirm_status": "Pending",
        "writeback_status": "Not Started",
        "version_count": 1,
        "item_count": 0,
        "import_remark": _clean_payload_text(payload, "import_remark") or "前端手工新增报关运单",
        "source_remark": _clean_payload_text(payload, "source_remark"),
    }
    return values


def create_batch(batch_payload: str | dict | None = None) -> dict:
    """新增一个空批次和默认当前版本，供后续手工加物料或文件补数。"""

    try:
        payload = _load_batch_payload(batch_payload)
    except Exception as exc:
        return {
            "ok": False,
            "dry_run": frappe is None,
            "message": f"新增报关运单参数解析失败：{exc}",
        }

    values = _build_manual_batch_values(payload)
    if not values["batch_no"]:
        return {
            "ok": False,
            "dry_run": frappe is None,
            "message": "请填写批次号/来源单号。",
        }

    version_values = {
        "version_code": f"手工-{values['batch_no']}",
        "version_type": "Estimated",
        "status": "Active",
        "is_current": 1,
        "source_type": "Manual",
        "fx_usd_to_rmb": DEFAULT_FX_USD_TO_RMB,
        "fx_rmb_to_mxn": DEFAULT_FX_RMB_TO_MXN,
        "remark": "前端手工新增批次默认版本",
    }

    if frappe is None:
        return {
            "ok": True,
            "dry_run": True,
            "batch": values,
            "version": version_values,
            "message": "当前未连接 Frappe，已返回新增报关运单预览。",
        }

    existing_name = frappe.db.get_value("Overseas Cost Batch", {"batch_no": values["batch_no"]}, "name")
    if existing_name:
        return {
            "ok": False,
            "batch_name": existing_name,
            "message": f"批次号已存在：{values['batch_no']}。请直接查询或换一个批次号。",
        }

    batch_doc = frappe.get_doc({"doctype": "Overseas Cost Batch", **values}).insert(ignore_permissions=True)
    version_doc = frappe.get_doc(
        {
            "doctype": "Overseas Cost Version",
            "batch": batch_doc.name,
            **version_values,
        }
    ).insert(ignore_permissions=True)
    frappe.db.set_value(
        "Overseas Cost Batch",
        batch_doc.name,
        {
            "current_version": version_doc.name,
            "version_count": 1,
            "item_count": 0,
        },
        update_modified=False,
    )
    frappe.get_doc(
        {
            "doctype": "Overseas Cost Audit Log",
            "batch": batch_doc.name,
            "version": version_doc.name,
            "action_type": "BATCH_EDIT",
            "field_name": "batch",
            "new_value": _json.dumps(values, ensure_ascii=False, default=str),
            "operator_name": getattr(frappe.session, "user", "") if getattr(frappe, "session", None) else "",
            "action_remark": "前端手工新增报关运单",
        }
    ).insert(ignore_permissions=True)
    frappe.db.commit()

    return {
        "ok": True,
        "batch_name": batch_doc.name,
        "version_name": version_doc.name,
        "batch_no": values["batch_no"],
        "message": "报关运单已新增，可继续添加物料或导入附件补数。",
    }


def _resolve_batch_name(batch_name: str) -> str | None:
    if frappe is None:
        return None
    batch = frappe.db.get_value("Overseas Cost Batch", batch_name, ["name"], as_dict=True)
    if batch:
        return batch["name"]
    batch = frappe.db.get_value("Overseas Cost Batch", {"batch_no": batch_name}, ["name"], as_dict=True)
    if batch:
        return batch["name"]
    return None


def _resolve_version_name(batch_doc_name: str, version_name: str | None = None) -> str | None:
    if frappe is None:
        return version_name
    if version_name:
        return version_name
    current_version = frappe.db.get_value("Overseas Cost Batch", batch_doc_name, "current_version")
    if current_version:
        return current_version
    rows = frappe.get_all(
        "Overseas Cost Version",
        filters={"batch": batch_doc_name},
        fields=["name"],
        order_by="modified desc",
        limit_page_length=1,
    )
    return rows[0]["name"] if rows else None


def _load_json(text: str | None) -> dict:
    if not text:
        return {}
    try:
        return _json.loads(text)
    except Exception:
        return {}


def _clean_query_value(value) -> str:
    if value in (None, ""):
        return ""
    return str(value).strip()


def _normalize_item_query_filters(filters: dict | None = None, **overrides) -> dict:
    query_filters = {}
    if isinstance(filters, dict):
        query_filters.update(filters)
    for key, value in overrides.items():
        if value not in (None, ""):
            query_filters[key] = value
    return {
        key: _clean_query_value(value)
        for key, value in query_filters.items()
        if _clean_query_value(value)
    }


def _build_item_query_args(
    batch_doc_name: str,
    version_name: str,
    filters: dict | None = None,
    *,
    keyword: str | None = None,
) -> tuple[list[list[str]], list[list[str]]]:
    query_filters = _normalize_item_query_filters(filters)
    keyword_value = _clean_query_value(keyword or query_filters.pop("keyword", ""))
    db_filters = [["batch", "=", batch_doc_name], ["version", "=", version_name]]

    for fieldname in ITEM_FILTER_FIELDS:
        value = query_filters.get(fieldname)
        if value:
            db_filters.append([fieldname, "like", f"%{value}%"])

    or_filters = []
    if keyword_value:
        or_filters = [[fieldname, "like", f"%{keyword_value}%"] for fieldname in ITEM_KEYWORD_FIELDS]

    return db_filters, or_filters


def _has_source_value(value) -> bool:
    return str(value or "").strip() != ""


def _source_status_key(value) -> str:
    return str(value or "").strip().lower()


def _is_parsed_attachment(row: dict) -> bool:
    return _source_status_key(row.get("parse_status")) == "parsed"


def _json_loads_dict(value) -> dict:
    if isinstance(value, dict):
        return value
    try:
        payload = json.loads(value or "{}")
    except (TypeError, json.JSONDecodeError):
        return {}
    return payload if isinstance(payload, dict) else {}


def _latest_tax_certificate_reconciliation(attachment_rows: list[dict]) -> dict:
    tax_rows = [
        row
        for row in attachment_rows
        if row.get("attachment_type") == "Tax Certificate" or row.get("source_type") == "Voucher"
    ]
    if not tax_rows:
        return {}

    sorted_rows = sorted(
        tax_rows,
        key=lambda row: str(row.get("modified") or row.get("creation") or ""),
        reverse=True,
    )
    for row in sorted_rows:
        mapped_result = _json_loads_dict(row.get("mapped_result_json"))
        if not mapped_result:
            continue
        parse_result = _json_loads_dict(row.get("parse_result_json"))
        summary = parse_result.get("summary") if isinstance(parse_result.get("summary"), dict) else {}
        voucher = mapped_result.get("voucher") if isinstance(mapped_result.get("voucher"), dict) else {}
        system = mapped_result.get("system") if isinstance(mapped_result.get("system"), dict) else {}
        difference = mapped_result.get("difference") if isinstance(mapped_result.get("difference"), dict) else {}
        manual_resolution = (
            mapped_result.get("manual_resolution")
            if isinstance(mapped_result.get("manual_resolution"), dict)
            else {}
        )
        return {
            "name": row.get("name") or "",
            "file_name": row.get("file_name") or "",
            "modified": row.get("modified") or "",
            "status": mapped_result.get("status") or "",
            "status_label": mapped_result.get("status_label") or "",
            "manual_resolution_status_label": manual_resolution.get("status_label") or "",
            "paid_total_mxn": (
                voucher.get("paid_total_mxn")
                if voucher.get("paid_total_mxn") is not None
                else summary.get("paid_total_mxn")
            ),
            "system_tax_total_mxn": system.get("system_import_tax_total_mxn"),
            "tax_total_diff_mxn": difference.get("tax_total_diff_mxn"),
            "direction_label": difference.get("direction_label") or "",
        }
    return {}


def _get_oa_logistics_trace(extra_json) -> dict:
    if isinstance(extra_json, dict):
        payload = dict(extra_json)
    else:
        try:
            payload = json.loads(extra_json or "{}")
        except (TypeError, json.JSONDecodeError):
            payload = {}
    if not isinstance(payload, dict):
        return {}
    trace = payload.get("oa_logistics_trace")
    return trace if isinstance(trace, dict) else payload


def _quote_candidate_summary(candidate: dict) -> dict:
    return {
        "carrier": str(candidate.get("carrier") or "").strip(),
        "amount": candidate.get("amount"),
        "currency": str(candidate.get("currency") or "").strip(),
        "volume_m3": candidate.get("volume_m3"),
        "gross_weight_kg": candidate.get("gross_weight_kg"),
        "chargeable_weight_kg": candidate.get("chargeable_weight_kg"),
        "unit_freight_per_kg": candidate.get("unit_freight_per_kg"),
        "billing_method": str(candidate.get("billing_method") or "").strip(),
        "allocation_basis": str(candidate.get("allocation_basis") or "").strip(),
        "pre_delivery_date": str(candidate.get("pre_delivery_date") or "").strip(),
        "destination": str(candidate.get("destination") or "").strip(),
        "source": str(candidate.get("source") or "").strip(),
        "evidence_line": str(candidate.get("evidence_line") or "").strip(),
        "source_field": str(candidate.get("source_field") or "").strip(),
        "status": str(candidate.get("status") or "待确认").strip(),
    }


def _public_logistics_text_summary(summary: dict) -> dict:
    if not isinstance(summary, dict):
        return {}
    allowed_fields = (
        "transport_mode",
        "transport_mode_raw",
        "logistics_no",
        "pre_delivery_date",
        "destination",
        "gross_weight_kg",
        "logistics_quote_amount",
        "logistics_quote_currency",
        "logistics_quote_carrier",
        "ai_used",
        "ai_model",
        "ai_confidence",
        "ai_reason",
    )
    public_summary = {
        key: summary.get(key)
        for key in allowed_fields
        if summary.get(key) not in (None, "")
    }
    evidence = str(summary.get("logistics_quote_evidence") or "").strip()
    if evidence:
        public_summary["logistics_quote_evidence"] = evidence[:180]
    return public_summary


def _logistics_text_summary(trace: dict) -> dict:
    summary = trace.get("logistics_text_summary")
    if isinstance(summary, dict):
        return _public_logistics_text_summary(summary)
    if isinstance(trace.get("form_fields"), dict):
        try:
            from overseas_costing.scripts.import_oa_logistics import extract_logistics_text_summary_from_approval

            return _public_logistics_text_summary(extract_logistics_text_summary_from_approval(trace))
        except Exception:
            return {}
    return {}


def _build_batch_source_status(batch: dict, attachments: list[dict] | None = None) -> dict:
    """按业务资料链路汇总批次来源状态，供前端数据检查展示。"""

    attachment_rows = attachments or []
    source_no = (
        batch.get("source_approval_no")
        or batch.get("source_instance_id")
        or batch.get("batch_no")
        or batch.get("name")
        or ""
    )
    has_oa_logistics = (
        batch.get("source_type") == "oa_logistics"
        or _has_source_value(batch.get("source_approval_no"))
        or _has_source_value(batch.get("source_instance_id"))
        or _has_source_value(batch.get("source_dingtalk_url"))
    )

    oa_attachment_rows = [row for row in attachment_rows if row.get("source_type") == "OA"]
    packing_list_rows = [row for row in attachment_rows if row.get("attachment_type") == "Packing List"]
    tax_certificate_rows = [
        row
        for row in attachment_rows
        if row.get("attachment_type") == "Tax Certificate" or row.get("source_type") == "Voucher"
    ]
    batch_source_attachment_count = int(batch.get("source_attachment_count") or 0)
    oa_attachment_count = max(batch_source_attachment_count, len(oa_attachment_rows))
    trace = _get_oa_logistics_trace(batch.get("extra_json"))
    quote_candidates = trace.get("logistics_quote_candidates") or []
    if not isinstance(quote_candidates, list):
        quote_candidates = []
    if not quote_candidates and isinstance(trace.get("form_fields"), dict):
        try:
            from overseas_costing.scripts.import_oa_logistics import extract_logistics_quote_candidates_from_approval

            quote_candidates = extract_logistics_quote_candidates_from_approval({"form_fields": trace["form_fields"]})
        except Exception:
            quote_candidates = []
    quote_candidates = [_quote_candidate_summary(row) for row in quote_candidates if isinstance(row, dict)]
    confirmed_quote = trace.get("confirmed_logistics_quote")
    confirmed_quote = _quote_candidate_summary(confirmed_quote) if isinstance(confirmed_quote, dict) else {}
    logistics_text_summary = _logistics_text_summary(trace)

    return {
        "source_no": source_no,
        "has_oa_logistics": bool(has_oa_logistics),
        "source_approval_status": batch.get("source_approval_status") or "",
        "oa_attachment_count": oa_attachment_count,
        "registered_attachment_count": len(attachment_rows),
        "packing_list_count": len(packing_list_rows),
        "parsed_packing_list_count": sum(1 for row in packing_list_rows if _is_parsed_attachment(row)),
        "tax_certificate_count": len(tax_certificate_rows),
        "parsed_tax_certificate_count": sum(1 for row in tax_certificate_rows if _is_parsed_attachment(row)),
        "has_form_attachments": oa_attachment_count > 0,
        "has_packing_list": bool(packing_list_rows),
        "has_tax_certificate": bool(tax_certificate_rows),
        "logistics_quote_candidate_count": len(quote_candidates),
        "logistics_quote_candidates": quote_candidates,
        "confirmed_logistics_quote": confirmed_quote,
        "logistics_text_summary": logistics_text_summary,
        "has_confirmed_logistics_quote": bool(confirmed_quote.get("amount")),
        "latest_tax_certificate_reconciliation": _latest_tax_certificate_reconciliation(attachment_rows),
        "source_priority_policy": source_priority_service.get_source_priority_policy(),
        "source_priority_summary": source_priority_service.get_source_priority_summary(),
    }


def _attach_batch_source_status(items: list[dict]) -> list[dict]:
    if not items:
        return items

    batch_names = [item.get("name") for item in items if item.get("name")]
    attachment_rows: list[dict] = []
    if batch_names:
        try:
            attachment_rows = frappe.get_all(
                "Overseas Cost Attachment",
                filters={"batch": ["in", batch_names]},
                fields=[
                    "name",
                    "batch",
                    "source_type",
                    "attachment_type",
                    "parse_status",
                    "file_name",
                    "parse_result_json",
                    "mapped_result_json",
                    "modified",
                ],
                limit_page_length=10000,
            )
        except Exception:
            attachment_rows = []

    attachments_by_batch: dict[str, list[dict]] = {}
    for row in attachment_rows:
        batch_name = row.get("batch")
        if not batch_name:
            continue
        attachments_by_batch.setdefault(batch_name, []).append(row)

    for item in items:
        item["source_status"] = _build_batch_source_status(
            item,
            attachments_by_batch.get(item.get("name"), []),
        )
    return items


def _attach_batch_calculation_snapshot(items: list[dict]) -> list[dict]:
    """把当前版本的 AI/分摊快照附加到批次列表，供右下角复核面板展示。"""

    if not items:
        return items

    version_names = [item.get("current_version") for item in items if item.get("current_version")]
    if not version_names:
        return items

    versions = frappe.get_all(
        "Overseas Cost Version",
        filters={"name": ["in", version_names]},
        fields=["name", "summary_snapshot_json", "rule_snapshot_json", "calculated_at"],
        limit_page_length=len(version_names),
    )
    versions_by_name = {version["name"]: version for version in versions}
    for item in items:
        version = versions_by_name.get(item.get("current_version")) or {}
        summary = _load_json(version.get("summary_snapshot_json"))
        rules = _load_json(version.get("rule_snapshot_json"))
        item["summary_snapshot"] = summary
        item["ai_allocation"] = summary.get("ai_allocation") or {}
        item["allocation_rule_snapshot"] = rules if isinstance(rules, list) else []
        item["calculated_at"] = version.get("calculated_at")
    return items


def _to_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def _to_int(value, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _classic_history_or_filters() -> list[list[str]]:
    or_filters: list[list[str]] = []
    for key in CLASSIC_HISTORY_BATCH_KEYS:
        or_filters.extend(
            [
                ["name", "=", key],
                ["batch_no", "=", key],
                ["source_approval_no", "=", key],
            ]
        )
    return or_filters


def _recent_start(recent_days: int) -> str:
    start = datetime.now() - timedelta(days=max(recent_days, 1))
    return start.strftime("%Y-%m-%d %H:%M:%S")


def _build_default_batch_time_filters(filters: dict) -> tuple[list, list, int, bool]:
    recent_days = max(_to_int(filters.get("recent_days"), DEFAULT_BATCH_RECENT_DAYS), 1)
    include_history = _to_bool(filters.get("include_history"))
    has_keyword = bool(str(filters.get("keyword") or "").strip())
    if include_history or has_keyword:
        return [], [], recent_days, False

    return [["source_created_at", ">=", _recent_start(recent_days)]], [], recent_days, True


def _dedupe_batches(rows: list[dict]) -> list[dict]:
    seen: set[str] = set()
    items: list[dict] = []
    for row in rows:
        name = row.get("name")
        if not name or name in seen:
            continue
        seen.add(name)
        items.append(row)
    return items


def _keyword_item_batch_names(keyword: str) -> list[str]:
    if not keyword:
        return []
    like_keyword = f"%{keyword}%"
    rows = frappe.get_all(
        "Overseas Cost Item",
        or_filters=[
            ["material_code", "like", like_keyword],
            ["product_name", "like", like_keyword],
            ["product_name_es", "like", like_keyword],
            ["import_name", "like", like_keyword],
            ["hs_code", "like", like_keyword],
            ["category", "like", like_keyword],
            ["customs_no", "like", like_keyword],
            ["waybill_no", "like", like_keyword],
            ["source_doc_no", "like", like_keyword],
        ],
        fields=["batch"],
        limit_page_length=200,
    )
    names: list[str] = []
    seen: set[str] = set()
    for row in rows:
        name = row.get("batch")
        if name and name not in seen:
            seen.add(name)
            names.append(name)
    return names


def get_batch_list(filters: dict) -> dict:
    if frappe is None:
        return {
            "ok": True,
            "dry_run": True,
            "message": "当前未连接 Frappe，返回空批次列表。",
            "filters": filters,
            "items": [],
            "total": 0,
        }

    db_filters = []
    if filters.get("transport_mode"):
        db_filters.append(["transport_mode", "=", filters["transport_mode"]])
    if filters.get("status"):
        db_filters.append(["status", "=", filters["status"]])
    recent_filters, default_or_filters, recent_days, recent_only = _build_default_batch_time_filters(filters)
    db_filters.extend(recent_filters)

    keyword = filters.get("keyword")
    fields = [
        "name",
        "batch_no",
        "customs_no",
        "waybill_no",
        "transport_mode",
        "project_collection",
        "source_type",
        "source_file_name",
        "source_sheet",
        "source_range",
        "source_approval_no",
        "source_instance_id",
        "source_dingtalk_url",
        "source_approval_status",
        "source_attachment_count",
        "source_created_at",
        "status",
        "current_version",
        "item_count",
        "total_goods_value",
        "total_gross_weight_kg",
        "estimated_total_cost_rmb",
        "actual_total_cost_rmb",
        "extra_json",
        "modified",
    ]
    query_kwargs = {
        "filters": db_filters,
        "fields": fields,
        "order_by": "source_created_at desc, modified desc",
        "limit_page_length": 200,
    }
    if keyword:
        like_keyword = f"%{keyword}%"
        query_kwargs["or_filters"] = [
            ["batch_no", "like", like_keyword],
            ["customs_no", "like", like_keyword],
            ["waybill_no", "like", like_keyword],
            ["source_approval_no", "like", like_keyword],
            ["source_instance_id", "like", like_keyword],
            ["project_collection", "like", like_keyword],
        ]

    items = frappe.get_all("Overseas Cost Batch", **query_kwargs)
    if keyword:
        item_batch_names = _keyword_item_batch_names(str(keyword).strip())
        if item_batch_names:
            item_filters = []
            if filters.get("transport_mode"):
                item_filters.append(["transport_mode", "=", filters["transport_mode"]])
            if filters.get("status"):
                item_filters.append(["status", "=", filters["status"]])
            item_filters.append(["name", "in", item_batch_names])
            item_batches = frappe.get_all(
                "Overseas Cost Batch",
                filters=item_filters,
                fields=fields,
                order_by="source_created_at desc, modified desc",
                limit_page_length=200,
            )
            items = _dedupe_batches(items + item_batches)
    if default_or_filters:
        classic_filters = []
        if filters.get("transport_mode"):
            classic_filters.append(["transport_mode", "=", filters["transport_mode"]])
        if filters.get("status"):
            classic_filters.append(["status", "=", filters["status"]])
        classic_items = frappe.get_all(
            "Overseas Cost Batch",
            filters=classic_filters,
            or_filters=default_or_filters,
            fields=fields,
            order_by="source_created_at desc, modified desc",
            limit_page_length=50,
        )
        for item in classic_items:
            item["is_classic_sample"] = 1
            item["sample_note"] = "历史样本"
        items = _dedupe_batches(items + classic_items)
        items.sort(key=lambda row: (str(row.get("source_created_at") or ""), str(row.get("modified") or "")), reverse=True)

    classic_keys = set(CLASSIC_HISTORY_BATCH_KEYS)
    for item in items:
        if item.get("batch_no") in classic_keys or item.get("source_approval_no") in classic_keys or item.get("name") in classic_keys:
            item["is_classic_sample"] = 1
            item["sample_note"] = "历史样本"

    items = [item for item in items if not is_hidden_approval_status(item.get("source_approval_status"))]
    items = _attach_batch_source_status(items)
    items = _attach_batch_calculation_snapshot(items)
    for item in items:
        item.pop("extra_json", None)
    return {
        "ok": True,
        "message": "批次列表已返回。",
        "filters": filters,
        "items": items,
        "total": len(items),
    }


def get_batch_detail(batch_name: str, version_name: str | None = None) -> dict:
    if frappe is None:
        source_meta = _get_batch_source_meta(batch_name)
        return {
            "ok": True,
            "dry_run": True,
            "message": "当前未连接 Frappe，返回批次详情预览。",
            "batch_name": batch_name,
            "version_name": version_name,
            "header": {"batch_no": source_meta.get("batch_no") or batch_name},
            "summary": {},
        }

    batch_doc_name = _resolve_batch_name(batch_name)
    if not batch_doc_name:
        return {"ok": False, "message": f"未找到批次：{batch_name}"}

    header = frappe.db.get_value(
        "Overseas Cost Batch",
        batch_doc_name,
        [
            "name",
            "batch_no",
            "customs_no",
            "waybill_no",
            "container_no",
            "sea_bill_no",
            "commercial_invoice_no",
            "transport_mode",
            "project_collection",
            "source_type",
            "source_file_name",
            "source_sheet",
            "source_range",
            "source_approval_no",
            "source_instance_id",
            "source_dingtalk_url",
            "status",
            "current_version",
            "confirm_status",
            "writeback_status",
            "item_count",
            "total_goods_value",
            "total_gross_weight_kg",
            "estimated_total_cost_rmb",
            "actual_total_cost_rmb",
        ],
        as_dict=True,
    )
    resolved_version_name = _resolve_version_name(batch_doc_name, version_name)
    version = {}
    summary = {}
    if resolved_version_name:
        version = frappe.db.get_value(
            "Overseas Cost Version",
            resolved_version_name,
            [
                "name",
                "version_code",
                "version_type",
                "status",
                "is_current",
                "fx_usd_to_rmb",
                "fx_rmb_to_mxn",
                "calculated_at",
                "summary_snapshot_json",
            ],
            as_dict=True,
        ) or {}
        summary = _load_json(version.get("summary_snapshot_json"))

    rules = []
    if resolved_version_name:
        rules = frappe.get_all(
            "Overseas Cost Allocation Rule",
            filters={"batch": batch_doc_name, "version": resolved_version_name},
            fields=["name", "rule_code", "expense_category", "allocation_basis", "currency", "amount", "remark", "is_enabled"],
            order_by="priority_no asc, modified asc",
            limit_page_length=1000,
        )

    return {
        "ok": True,
        "message": "批次详情已返回。",
        "batch_name": batch_doc_name,
        "version_name": resolved_version_name,
        "header": header,
        "version": version,
        "summary": summary,
        "allocation_rules": rules,
    }


def get_batch_items(
    batch_name: str,
    version_name: str | None = None,
    filters: dict | None = None,
    keyword: str | None = None,
    customs_no: str | None = None,
    waybill_no: str | None = None,
    material_code: str | None = None,
    product_name: str | None = None,
    import_name: str | None = None,
    hs_code: str | None = None,
    category: str | None = None,
) -> dict:
    query_filters = _normalize_item_query_filters(
        filters,
        keyword=keyword,
        customs_no=customs_no,
        waybill_no=waybill_no,
        material_code=material_code,
        product_name=product_name,
        import_name=import_name,
        hs_code=hs_code,
        category=category,
    )
    if frappe is None:
        return {
            "ok": True,
            "dry_run": True,
            "message": "当前未连接 Frappe，返回空明细。",
            "batch_name": batch_name,
            "version_name": version_name,
            "filters": query_filters,
            "columns": EXCEL_COLUMNS,
            "items": [],
        }

    batch_doc_name = _resolve_batch_name(batch_name)
    if not batch_doc_name:
        return {"ok": False, "message": f"未找到批次：{batch_name}"}

    resolved_version_name = _resolve_version_name(batch_doc_name, version_name)
    if not resolved_version_name:
        return {"ok": False, "batch_name": batch_doc_name, "message": "当前批次没有版本。"}

    fields = list(dict.fromkeys(EXTRA_ITEM_FIELDS + EXCEL_FIELDNAMES))
    db_filters, or_filters = _build_item_query_args(
        batch_doc_name,
        resolved_version_name,
        query_filters,
    )
    items = frappe.get_all(
        "Overseas Cost Item",
        filters=db_filters,
        or_filters=or_filters,
        fields=fields,
        order_by="row_no asc",
        limit_page_length=10000,
    )
    return {
        "ok": True,
        "message": "批次明细已返回。",
        "batch_name": batch_doc_name,
        "version_name": resolved_version_name,
        "filters": query_filters,
        "columns": EXCEL_COLUMNS,
        "items": items,
        "total": len(items),
    }


def get_version_list(batch_name: str) -> dict:
    if frappe is None:
        return {
            "ok": True,
            "dry_run": True,
            "message": "当前未连接 Frappe，返回空版本列表。",
            "batch_name": batch_name,
            "items": [],
        }

    batch_doc_name = _resolve_batch_name(batch_name)
    if not batch_doc_name:
        return {"ok": False, "message": f"未找到批次：{batch_name}"}

    items = frappe.get_all(
        "Overseas Cost Version",
        filters={"batch": batch_doc_name},
        fields=[
            "name",
            "version_code",
            "version_type",
            "status",
            "is_current",
            "source_type",
            "fx_usd_to_rmb",
            "fx_rmb_to_mxn",
            "calculated_at",
            "modified",
        ],
        order_by="creation asc",
        limit_page_length=1000,
    )
    return {
        "ok": True,
        "message": "版本列表已返回。",
        "batch_name": batch_doc_name,
        "items": items,
    }


def _normalize_limit(limit, default: int = 80, maximum: int = 300) -> int:
    try:
        value = int(limit)
    except (TypeError, ValueError):
        value = default
    return max(1, min(value, maximum))


EXPORT_XLSX_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
EXPORT_TEXT_FIELDS = {
    "material_code",
    "product_name",
    "product_name_es",
    "spec_model",
    "unit",
    "recipient",
    "purchase_currency",
    "import_name",
    "hs_code",
    "category",
    "customs_no",
    "waybill_no",
    "container_no",
    "sea_bill_no",
    "commercial_invoice_no",
    "purchase_order_no",
    "project_collection",
    "transport_mode",
    "source_type",
    "source_doc_no",
    "source_file_name",
}
EXPORT_BATCH_FALLBACK_FIELDS = [
    "name",
    "batch_no",
    "customs_no",
    "waybill_no",
    "container_no",
    "sea_bill_no",
    "commercial_invoice_no",
    "transport_mode",
    "project_collection",
    "source_approval_no",
    "source_instance_id",
    "source_dingtalk_url",
    "source_file_name",
    "current_version",
]
TRANSPORT_MODE_LABELS = {"SEA": "海运", "AIR": "空运", "EXPRESS": "快递"}
EXPORT_MERGE_FIELDNAMES = {
    "customs_no",
    "waybill_no",
    "china_misc_rmb",
    "china_misc_mxn",
    "china_ocean_usd",
    "china_to_mexico_freight_rmb",
    "project_collection",
    "transport_mode",
}
EXPORT_ZERO_AS_EMPTY_FIELDNAMES = {
    "limpieza_contenedor",
}
EXPORT_MIN_COLUMN_WIDTHS = {
    "material_code": 14,
    "product_name": 28,
    "spec_model": 28,
    "unit": 12,
    "import_name": 34,
    "hs_code": 16,
    "category": 20,
    "customs_no": 24,
    "waybill_no": 24,
    "project_collection": 36,
    "transport_mode": 14,
}
EXPORT_MAX_COLUMN_WIDTHS = {
    "product_name": 42,
    "spec_model": 42,
    "unit": 16,
    "import_name": 48,
    "category": 34,
    "customs_no": 30,
    "waybill_no": 30,
    "project_collection": 52,
    "transport_mode": 18,
}
EXPORT_WRAP_FIELDNAMES = {
    "product_name",
    "spec_model",
    "import_name",
    "category",
    "customs_no",
    "waybill_no",
    "project_collection",
    "transport_mode",
}
EXPORT_HEADER_LABELS = {
    "spec_model": "规格型号\nEspecificación / Modelo",
    "unit_price": "采购单价",
    "quantity": "采购数量\n按采购单位",
    "unit": "单位\nUnidad",
    "waybill_no": "中国到墨西哥\n运单号",
    "china_misc_rmb": "中国运输及\n相关杂费\nRMB",
    "china_misc_mxn": "中国运输及\n相关杂费\n折合MXN PESOS",
    "china_ocean_usd": "中国海运\n(USD)",
    "goods_value_ratio": "分摊比例\n(货值比）",
    "import_tax_total": "IMPUESTOS\n合计清关税费",
    "revalidacion": "REVALIDACION\n文件验证费",
    "maniobras": "MANIOBRAS\n码头操作费",
    "entrega_mercancia": "ENTREGA DE \nMERCANCIA\n配送费用",
    "previo": "PREVIO\n预检费用",
    "service_aa": "Servicios A.A.\n货代服务费",
    "complemento_maniobras": "COMPLEMENTO DE MANIOBRAS\n（操作辅助装置）",
    "desconsolidacion": "DESCONSOLIDACION\n拆箱分货费用",
    "maniobra_falso": "MANIOBRA EN FALSO\n虚假操作费",
    "limpieza_contenedor": "LIMPIEZA DE CONTENEDOR\n集装箱清洁",
    "mexico_customs_mxn": "墨西哥\n清关费用\n(MXN)",
    "mexico_customs_rmb": "墨西哥\n清关费用\n(RMB )",
    "mexico_customs_usd": "墨西哥\n清关费用\n（USD）",
    "mexico_inland_mxn": "墨西哥\n内陆运输\n费用\n（MXN）",
    "mexico_misc_mxn": "墨西哥杂费\n（MXN）",
    "mexico_inland_misc_rmb": "墨西哥\n内陆运输+杂费\n费用\n（RMB）",
    "china_to_mexico_freight_rmb": "中国到\n墨西哥运费\n（RMB）",
    "gross_weight_kg": "货重毛重\n（kg）",
    "weight_ratio": "分摊比例\n（重量比）",
    "freight_alloc_rmb": "中国到墨西哥\n运输费用分摊\nRMB",
    "freight_alloc_mxn": "中国到墨西哥\n运输费用分摊\nMXN PESOS",
    "total_logistics_mxn": "中国到墨西哥\n运输+清关+杂费\nMXN PESOS",
    "alloc_price_mxn": "分摊运输+物流+杂费\n到产品的价格\nMXN PESOS",
    "total_cost_rmb": "综合成本\nRMB",
    "total_unit_rmb": "综合单价\nRMB\n按采购单位",
}
EXPORT_HEADER_COLOR_GROUPS = {
    "tax": {
        "cc_rate",
        "cc_anti_dumping",
        "igi_rate",
        "igi_amount",
        "iva_rate",
        "iva_amount",
        "goods_value_ratio",
        "dta",
        "prv_duty",
        "prv_iva",
    },
    "customs_fee": {
        "import_tax_total",
        "revalidacion",
        "maniobras",
        "muellaje",
        "entrega_mercancia",
        "previo",
        "service_aa",
        "almacenajes",
        "reconocimiento_aduanero",
        "honorarios",
        "complemento_maniobras",
        "desconsolidacion",
        "maniobra_falso",
        "arrastre",
        "patio_regulador",
        "entrega_vacio",
        "limpieza_contenedor",
    },
    "mexico_customs": {"mexico_customs_mxn", "mexico_customs_rmb"},
    "allocation": {"gross_weight_kg", "weight_ratio", "freight_alloc_rmb", "freight_alloc_mxn"},
    "result": {"total_logistics_mxn", "alloc_price_mxn", "total_cost_rmb", "total_unit_rmb"},
}


def _normalize_export_batch_names(batch_names_json) -> list[str]:
    if isinstance(batch_names_json, str):
        text = batch_names_json.strip()
        if not text:
            return []
        try:
            value = json.loads(text)
        except Exception:
            value = [part.strip() for part in text.split(",")]
    else:
        value = batch_names_json
    if not isinstance(value, (list, tuple)):
        return []
    result = []
    for item in value:
        name = str(item or "").strip()
        if name and name not in result:
            result.append(name)
    return result


def _transport_label(value) -> str:
    code = normalize_transport_mode(value)
    return TRANSPORT_MODE_LABELS.get(code, str(value or ""))


def _clean_export_filename_part(value: str) -> str:
    text = str(value or "").strip() or "全部"
    for char in '\\/:*?"<>|':
        text = text.replace(char, "_")
    return text


def _export_cell_value(item: dict, batch: dict, column: dict):
    fieldname = column.get("fieldname")
    value = item.get(fieldname)
    if value in (None, ""):
        value = batch.get(fieldname)
    if fieldname == "transport_mode":
        value = _transport_label(value)
    if value in (None, ""):
        return ""
    if fieldname in EXPORT_ZERO_AS_EMPTY_FIELDNAMES:
        try:
            if float(value) == 0:
                return ""
        except (TypeError, ValueError):
            pass
    if fieldname in EXPORT_TEXT_FIELDS:
        return str(value)
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    return int(number) if number.is_integer() else number


def _display_width(value) -> int:
    text = str(value or "")
    return sum(2 if ord(char) > 127 else 1 for char in text)


def _export_column_width(fieldname: str, header: str, values: list) -> int:
    max_width = _display_width(header) + 2
    for value in values[:500]:
        max_width = max(max_width, _display_width(value) + 2)
    min_width = EXPORT_MIN_COLUMN_WIDTHS.get(fieldname, 12)
    max_limit = EXPORT_MAX_COLUMN_WIDTHS.get(fieldname, 40)
    return min(max(max_width, min_width), max_limit)


def _export_header_label(column: dict) -> str:
    fieldname = column.get("fieldname") or ""
    return EXPORT_HEADER_LABELS.get(fieldname) or str(column.get("label") or fieldname or "").strip()


def _export_header_fill_color(fieldname: str) -> str:
    if fieldname in EXPORT_HEADER_COLOR_GROUPS["tax"]:
        return "FF0DF2DF"
    if fieldname in EXPORT_HEADER_COLOR_GROUPS["customs_fee"]:
        return "FFD04FCA"
    if fieldname in EXPORT_HEADER_COLOR_GROUPS["mexico_customs"]:
        return "FFFCC102"
    if fieldname in EXPORT_HEADER_COLOR_GROUPS["allocation"]:
        return "FFFFF3CE"
    if fieldname in EXPORT_HEADER_COLOR_GROUPS["result"]:
        return "FF04B0F1"
    if fieldname in {"project_collection", "transport_mode"}:
        return "FFFFFFFF"
    return "FFFFD966"


def _export_header_font(fieldname: str):
    from openpyxl.styles import Font

    if fieldname in EXPORT_HEADER_COLOR_GROUPS["result"]:
        return Font(name="宋体", size=8, bold=True, color="FFFFFFFF")
    if fieldname in {"project_collection", "transport_mode"}:
        return Font(name="宋体", size=8, bold=True, color="FF000000")
    return Font(name="宋体", size=8, bold=False, color="FF000000")


def _merge_repeated_export_cells(sheet, columns: list[dict], rows: list[list], merge_alignment, cell_border) -> None:
    """把连续相同的整票字段纵向合并，保留 SKU 行级字段逐行展示。"""

    if len(rows) < 2:
        return

    merge_column_indexes = [
        index
        for index, column in enumerate(columns, start=1)
        if column.get("fieldname") in EXPORT_MERGE_FIELDNAMES
    ]
    if not merge_column_indexes:
        return

    for column_index in merge_column_indexes:
        start_row = 2
        previous_value = sheet.cell(row=start_row, column=column_index).value
        for row_index in range(3, len(rows) + 3):
            current_value = sheet.cell(row=row_index, column=column_index).value if row_index <= len(rows) + 1 else None
            if current_value == previous_value:
                continue
            if previous_value not in (None, "") and row_index - start_row > 1:
                sheet.merge_cells(
                    start_row=start_row,
                    start_column=column_index,
                    end_row=row_index - 1,
                    end_column=column_index,
                )
                merged_cell = sheet.cell(row=start_row, column=column_index)
                merged_cell.alignment = merge_alignment
                merged_cell.border = cell_border
            start_row = row_index
            previous_value = current_value


def _build_export_xlsx_content(columns: list[dict], rows: list[list]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except Exception as exc:  # pragma: no cover - 真实导出时才需要 openpyxl
        raise RuntimeError("导出 .xlsx 需要安装 openpyxl。") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "综合成本核算"
    sheet.freeze_panes = "C2"

    headers = [_export_header_label(column) for column in columns]
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    last_column_letter = get_column_letter(len(headers))
    sheet.auto_filter.ref = f"A1:{last_column_letter}{max(sheet.max_row, 1)}"

    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_font = Font(name="等线", size=10, bold=False, color="FF000000")
    small_body_font = Font(name="等线", size=8, bold=False, color="FF000000")
    body_alignment = Alignment(horizontal="center", vertical="center")
    center_wrap_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_wrap_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    merged_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side = Side(style="thin", color="FF000000")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    sheet.row_dimensions[1].height = 53.25
    for column_index, cell in enumerate(sheet[1], start=1):
        fieldname = columns[column_index - 1].get("fieldname") if column_index <= len(columns) else ""
        cell.fill = PatternFill(fill_type="solid", fgColor=_export_header_fill_color(fieldname))
        cell.font = _export_header_font(fieldname)
        cell.alignment = header_alignment
        cell.border = cell_border

    sheet.sheet_view.showGridLines = True
    for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        sheet.row_dimensions[row_index].height = 20
        for column_index, cell in enumerate(row, start=1):
            fieldname = columns[column_index - 1].get("fieldname") if column_index <= len(columns) else ""
            cell.font = small_body_font if fieldname == "import_name" else body_font
            if fieldname in {"product_name", "project_collection"}:
                cell.alignment = left_wrap_alignment
            elif fieldname in EXPORT_WRAP_FIELDNAMES:
                cell.alignment = center_wrap_alignment
            else:
                cell.alignment = body_alignment
            cell.border = cell_border
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.######'

    _merge_repeated_export_cells(sheet, columns, rows, merged_alignment, cell_border)

    for index, header in enumerate(headers, start=1):
        fieldname = columns[index - 1].get("fieldname") if index - 1 < len(columns) else ""
        values = [row[index - 1] if index - 1 < len(row) else "" for row in rows]
        sheet.column_dimensions[get_column_letter(index)].width = _export_column_width(fieldname, header, values)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def export_current_result_xlsx(batch_names_json=None, transport_label: str | None = None) -> dict:
    if frappe is None:
        return {
            "ok": False,
            "dry_run": True,
            "message": "当前未连接 Frappe，不能导出真实结果。",
        }

    batch_names = _normalize_export_batch_names(batch_names_json)
    if not batch_names:
        return {"ok": False, "message": "当前没有可导出的批次。"}

    export_rows = []
    for batch_name in batch_names:
        batch_doc_name = _resolve_batch_name(batch_name)
        if not batch_doc_name:
            continue
        batch = frappe.db.get_value(
            "Overseas Cost Batch",
            batch_doc_name,
            EXPORT_BATCH_FALLBACK_FIELDS,
            as_dict=True,
        ) or {}
        detail = get_batch_items(
            batch_name=batch_doc_name,
            version_name=batch.get("current_version"),
        )
        if not detail.get("ok"):
            continue
        for item in detail.get("items") or []:
            export_rows.append([_export_cell_value(item, batch, column) for column in EXCEL_COLUMNS])

    if not export_rows:
        return {"ok": False, "message": "当前批次没有可导出的 SKU 明细。"}

    content = _build_export_xlsx_content(EXCEL_COLUMNS, export_rows)
    label = _clean_export_filename_part(transport_label or "全部")
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"海外采购综合成本核算_{label}_{stamp}.xlsx"
    return {
        "ok": True,
        "file_name": file_name,
        "mime_type": EXPORT_XLSX_MIME_TYPE,
        "content_base64": base64.b64encode(content).decode("ascii"),
        "total": len(export_rows),
        "message": f"已生成 {len(export_rows)} 行 SKU 明细。",
    }


WRITEBACK_REQUIRED_ITEM_FIELDS = (
    ("material_code", "物料编码", "text"),
    ("product_name", "物料名称", "text"),
    ("quantity", "数量", "positive_number"),
    ("unit_price", "采购单价", "positive_number"),
    ("purchase_currency", "采购币种", "text"),
    ("goods_value", "总货值", "positive_number"),
    ("total_unit_rmb", "综合物品单价RMB", "positive_number"),
)


def _as_float(value) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _is_blank(value) -> bool:
    return value is None or str(value).strip() == ""


def _build_writeback_item_quality(items: list[dict]) -> dict:
    issue_counts = {fieldname: 0 for fieldname, _label, _rule in WRITEBACK_REQUIRED_ITEM_FIELDS}
    issue_examples = []

    for index, item in enumerate(items, start=1):
        item_missing_labels = []
        for fieldname, label, rule in WRITEBACK_REQUIRED_ITEM_FIELDS:
            value = item.get(fieldname)
            has_issue = _is_blank(value) if rule == "text" else _as_float(value) <= 0
            if has_issue:
                issue_counts[fieldname] += 1
                item_missing_labels.append(label)

        if item_missing_labels and len(issue_examples) < 5:
            issue_examples.append(
                {
                    "row_no": item.get("row_no") or item.get("excel_row_no") or index,
                    "material_code": item.get("material_code") or "",
                    "product_name": item.get("product_name") or "",
                    "missing_fields": item_missing_labels,
                }
            )

    checks = {
        f"items_have_{fieldname}": count == 0
        for fieldname, count in issue_counts.items()
    }
    blocking_reasons = [
        f"有 {count} 条 SKU 缺少或未填有效的{label}。"
        for fieldname, label, _rule in WRITEBACK_REQUIRED_ITEM_FIELDS
        if (count := issue_counts[fieldname]) > 0
    ]

    return {
        "checks": checks,
        "issue_counts": issue_counts,
        "issue_examples": issue_examples,
        "blocking_reasons": blocking_reasons,
    }


def _build_writeback_readiness(
    batch: dict,
    items: list[dict],
    resolved_version_name: str | None,
) -> dict:
    item_quality = _build_writeback_item_quality(items)
    actual_total_cost = _as_float(batch.get("actual_total_cost_rmb"))
    estimated_total_cost = _as_float(batch.get("estimated_total_cost_rmb"))
    total_cost = actual_total_cost or estimated_total_cost
    recorded_item_count = int(_as_float(batch.get("item_count")))
    actual_item_count = len(items)

    checks = {
        "batch_exists": True,
        "has_current_version": bool(resolved_version_name or batch.get("current_version")),
        "is_confirmed": batch.get("confirm_status") == "Confirmed",
        "has_dirty_data": batch.get("status") == "Dirty",
        "has_items": actual_item_count > 0,
        "has_total_cost": total_cost > 0,
        **item_quality["checks"],
    }

    blocking_reasons = []
    if not checks["has_current_version"]:
        blocking_reasons.append("当前批次没有当前版本。")
    if not checks["is_confirmed"]:
        blocking_reasons.append("当前批次还没有确认。")
    if checks["has_dirty_data"]:
        blocking_reasons.append("当前批次存在未重新计算的数据。")
    if not checks["has_items"]:
        blocking_reasons.append("当前批次没有 SKU 明细。")
    if not checks["has_total_cost"]:
        blocking_reasons.append("当前批次没有可回写的综合成本结果。")
    blocking_reasons.extend(item_quality["blocking_reasons"])

    warning_reasons = []
    if recorded_item_count and recorded_item_count != actual_item_count:
        warning_reasons.append(f"批次记录明细数为 {recorded_item_count}，实际查询到 {actual_item_count} 条。")
    if estimated_total_cost > 0 and actual_total_cost <= 0:
        warning_reasons.append("当前只有系统计算成本，尚无凭证后的实际总成本。")

    ready = not blocking_reasons
    return {
        "ready": ready,
        "checks": checks,
        "blocking_reasons": blocking_reasons,
        "warning_reasons": warning_reasons,
        "item_issue_counts": item_quality["issue_counts"],
        "item_issue_examples": item_quality["issue_examples"],
        "item_count": actual_item_count,
        "total_cost_rmb": total_cost,
        "message": "允许回写。" if ready else "当前批次暂不满足回写条件：" + "；".join(blocking_reasons),
    }


def get_audit_logs(batch_name: str, version_name: str | None = None, limit: int | str = 80) -> dict:
    normalized_limit = _normalize_limit(limit)
    if frappe is None:
        return {
            "ok": True,
            "dry_run": True,
            "message": "当前未连接 Frappe，返回空修改记录。",
            "batch_name": batch_name,
            "version_name": version_name,
            "items": [],
            "total": 0,
        }

    batch_doc_name = _resolve_batch_name(batch_name)
    if not batch_doc_name:
        return {"ok": False, "message": f"未找到批次：{batch_name}", "items": [], "total": 0}

    filters = {"batch": batch_doc_name}
    if version_name:
        filters["version"] = version_name

    items = frappe.get_all(
        "Overseas Cost Audit Log",
        filters=filters,
        fields=[
            "name",
            "batch",
            "version",
            "action_type",
            "field_name",
            "row_no",
            "old_value",
            "new_value",
            "operator_name",
            "action_remark",
            "creation",
        ],
        order_by="creation desc",
        limit_page_length=normalized_limit,
    )
    return {
        "ok": True,
        "message": "修改记录已返回。",
        "batch_name": batch_doc_name,
        "version_name": version_name,
        "items": items,
        "total": len(items),
    }


def check_writeback_ready(batch_name: str, version_name: str | None = None) -> dict:
    if frappe is None:
        blocking_reasons = ["当前未连接 Frappe，不能执行真实回写检查。"]
        return {
            "ok": True,
            "dry_run": True,
            "ready": False,
            "batch_name": batch_name,
            "version_name": version_name,
            "checks": {
                "batch_exists": False,
                "has_current_version": False,
                "is_confirmed": False,
                "has_dirty_data": False,
                "has_items": False,
                "has_total_cost": False,
            },
            "blocking_reasons": blocking_reasons,
            "warning_reasons": [],
            "item_issue_counts": {},
            "item_issue_examples": [],
            "item_count": 0,
            "total_cost_rmb": 0,
            "message": blocking_reasons[0],
        }

    batch_doc_name = _resolve_batch_name(batch_name)
    if not batch_doc_name:
        return {
            "ok": False,
            "ready": False,
            "batch_name": batch_name,
            "version_name": version_name,
            "checks": {"batch_exists": False},
            "blocking_reasons": [f"未找到批次：{batch_name}"],
            "warning_reasons": [],
            "message": f"未找到批次：{batch_name}",
        }

    resolved_version_name = _resolve_version_name(batch_doc_name, version_name)
    batch = frappe.db.get_value(
        "Overseas Cost Batch",
        batch_doc_name,
        [
            "status",
            "confirm_status",
            "current_version",
            "item_count",
            "estimated_total_cost_rmb",
            "actual_total_cost_rmb",
        ],
        as_dict=True,
    ) or {}
    item_filters = {"batch": batch_doc_name}
    if resolved_version_name:
        item_filters["version"] = resolved_version_name
    items = frappe.get_all(
        "Overseas Cost Item",
        filters=item_filters,
        fields=[
            "name",
            "row_no",
            "excel_row_no",
            "material_code",
            "product_name",
            "quantity",
            "unit_price",
            "purchase_currency",
            "goods_value",
            "total_unit_rmb",
        ],
        limit_page_length=10000,
    )
    readiness = _build_writeback_readiness(
        batch=batch,
        items=items,
        resolved_version_name=resolved_version_name,
    )
    return {
        "ok": True,
        "batch_name": batch_doc_name,
        "version_name": resolved_version_name,
        **readiness,
    }

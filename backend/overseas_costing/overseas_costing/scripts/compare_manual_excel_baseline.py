"""中文用途：生成系统试算结果与人工 Excel 核算结果的对照表。"""

from __future__ import annotations

import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from overseas_costing.services import batch_service

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover - 真实导出时才需要 openpyxl
    Workbook = None
    load_workbook = None
    Alignment = Border = Font = PatternFill = Side = None
    get_column_letter = None


DEFAULT_BATCH_NO = "HPCU5155607"
DEFAULT_CUSTOMS_NO = "26 16 1681 6000151"
DEFAULT_SOURCE_SHEET = "2026年YUEWEI"
DEFAULT_CANDIDATE_PATHS = [
    Path("/mnt/c/Users/lin/OneDrive/Desktop/墨西哥进口物料综合成本核算.xlsx"),
    Path("C:/Users/lin/OneDrive/Desktop/墨西哥进口物料综合成本核算.xlsx"),
    Path.cwd() / "data" / "墨西哥进口物料综合成本核算.xlsx",
]

MANUAL_COLUMNS = [
    ("manual_excel_row", None, "人工表行号"),
    ("material_code", "A", "物料编码"),
    ("product_name", "B", "产品名称"),
    ("unit_price", "C", "单价"),
    ("quantity", "D", "数量"),
    ("goods_value", "E", "总货值"),
    ("import_name", "F", "海关进口名称"),
    ("hs_code", "G", "海关分类编码"),
    ("category", "H", "大类分类"),
    ("customs_no", "I", "报关单号"),
    ("waybill_no", "J", "运单号"),
    ("china_misc_rmb", "K", "中国杂费RMB"),
    ("china_misc_mxn", "L", "中国杂费MXN"),
    ("china_ocean_usd", "M", "中国海运USD"),
    ("igi_amount", "Q", "IGI关税"),
    ("iva_amount", "S", "IVA增值税"),
    ("import_tax_total", "X", "清关税费合计"),
    ("mexico_customs_mxn", "AO", "墨西哥清关MXN"),
    ("mexico_customs_rmb", "AP", "墨西哥清关RMB"),
    ("mexico_customs_usd", "AQ", "墨西哥清关USD"),
    ("mexico_inland_mxn", "AR", "墨西哥内陆运输MXN"),
    ("mexico_misc_mxn", "AS", "墨西哥杂费MXN"),
    ("mexico_inland_misc_rmb", "AT", "墨西哥内陆运输+杂费RMB"),
    ("china_to_mexico_freight_rmb", "AU", "整票运费RMB"),
    ("gross_weight_kg", "AV", "毛重KG"),
    ("weight_ratio", "AW", "重量比"),
    ("freight_alloc_rmb", "AX", "运费分摊RMB"),
    ("freight_alloc_mxn", "AY", "运费分摊MXN"),
    ("total_logistics_mxn", "AZ", "运输清关杂费MXN"),
    ("alloc_price_mxn", "BA", "分摊到单价MXN"),
    ("total_cost_rmb", "BB", "综合成本RMB"),
    ("total_unit_rmb", "BC", "综合物品单价RMB"),
    ("project_collection", "BD", "项目归集"),
    ("transport_mode", "BE", "运输方式"),
]

COMPARE_FIELDS = [
    ("goods_value", "货值RMB"),
    ("gross_weight_kg", "毛重KG"),
    ("import_tax_total", "清关税费MXN"),
    ("mexico_customs_rmb", "清关费用RMB"),
    ("mexico_inland_misc_rmb", "墨西哥内陆/杂费RMB"),
    ("freight_alloc_rmb", "运费分摊RMB"),
    ("total_logistics_mxn", "物流清关合计MXN"),
    ("total_cost_rmb", "综合成本RMB"),
    ("total_unit_rmb", "综合单价RMB"),
]


def build_hpcu_manual_comparison_from_env() -> dict:
    return build_hpcu_manual_comparison(
        file_path=os.environ.get("OVERSEAS_COST_MANUAL_EXCEL"),
        batch_name=os.environ.get("OVERSEAS_COST_BATCH") or DEFAULT_BATCH_NO,
        customs_no=os.environ.get("OVERSEAS_COST_CUSTOMS_NO") or DEFAULT_CUSTOMS_NO,
        waybill_no=os.environ.get("OVERSEAS_COST_WAYBILL_NO") or DEFAULT_BATCH_NO,
        source_sheet=os.environ.get("OVERSEAS_COST_SOURCE_SHEET") or DEFAULT_SOURCE_SHEET,
        output_path=os.environ.get("OVERSEAS_COST_OUTPUT"),
    )


def build_hpcu_manual_comparison(
    *,
    file_path: str | None = None,
    batch_name: str = DEFAULT_BATCH_NO,
    customs_no: str = DEFAULT_CUSTOMS_NO,
    waybill_no: str = DEFAULT_BATCH_NO,
    source_sheet: str = DEFAULT_SOURCE_SHEET,
    output_path: str | None = None,
) -> dict:
    manual_path = _resolve_manual_workbook(file_path)
    manual = load_manual_baseline(
        file_path=manual_path,
        source_sheet=source_sheet,
        customs_no=customs_no,
        waybill_no=waybill_no,
    )
    system_result = batch_service.get_batch_items(batch_name=batch_name)
    if not system_result.get("ok"):
        return {"ok": False, "message": system_result.get("message") or "系统批次明细读取失败。"}

    system_rows = system_result.get("items") or []
    system_export = {
        **system_result,
        "business_batch_no": batch_name,
        "batch_doc_name": system_result.get("batch_name"),
    }
    comparison = build_comparison_rows(manual["items"], system_rows)
    summary_rows = build_summary_rows(manual["items"], system_rows)
    output = Path(output_path).expanduser() if output_path else _default_output_path(waybill_no)
    write_comparison_xlsx(
        output,
        manual=manual,
        system_result=system_export,
        summary_rows=summary_rows,
        comparison_rows=comparison["rows"],
        unmatched_system_rows=comparison["unmatched_system_rows"],
    )
    return {
        "ok": True,
        "file_path": str(manual_path),
        "output_path": str(output),
        "batch_name": batch_name,
        "batch_doc_name": system_result.get("batch_name"),
        "version_name": system_result.get("version_name"),
        "manual_item_count": len(manual["items"]),
        "system_item_count": len(system_rows),
        "matched_count": comparison["matched_count"],
        "manual_unmatched_count": comparison["manual_unmatched_count"],
        "system_unmatched_count": len(comparison["unmatched_system_rows"]),
        "summary_rows": summary_rows,
        "message": f"已生成系统与人工核算对照表：{output}",
    }


def load_manual_baseline(
    *,
    file_path: str | Path,
    source_sheet: str = DEFAULT_SOURCE_SHEET,
    customs_no: str = DEFAULT_CUSTOMS_NO,
    waybill_no: str = DEFAULT_BATCH_NO,
) -> dict:
    if load_workbook is None:
        raise RuntimeError("读取人工 Excel 需要安装 openpyxl。")

    workbook = load_workbook(file_path, data_only=True, read_only=False)
    try:
        worksheet = workbook[source_sheet] if source_sheet in workbook.sheetnames else _find_sheet_with_token(workbook, customs_no, waybill_no)
        start_row = _find_start_row(worksheet, customs_no, waybill_no)
        if not start_row:
            raise ValueError(f"人工 Excel 中未找到报关单/运单：{customs_no} / {waybill_no}")
        rows = _read_manual_block(worksheet, start_row, customs_no, waybill_no)
        return {
            "source_file": Path(file_path).name,
            "source_sheet": worksheet.title,
            "source_range": f"{worksheet.title}!{rows[0]['manual_excel_row']}:{rows[-1]['manual_excel_row']}" if rows else "",
            "customs_no": customs_no,
            "waybill_no": waybill_no,
            "items": rows,
        }
    finally:
        workbook.close()


def build_comparison_rows(manual_rows: list[dict], system_rows: list[dict]) -> dict:
    system_pool: list[dict] = [dict(row) for row in system_rows]
    comparison_rows = []
    matched_count = 0

    for manual in manual_rows:
        system = _pop_best_system_match(manual, system_pool)
        if system:
            matched_count += 1
        comparison_rows.append(_build_comparison_row(manual, system))

    return {
        "rows": comparison_rows,
        "matched_count": matched_count,
        "manual_unmatched_count": len(comparison_rows) - matched_count,
        "unmatched_system_rows": system_pool,
    }


def build_summary_rows(manual_rows: list[dict], system_rows: list[dict]) -> list[dict]:
    rows = [{"field": "item_count", "label": "物料行数", "manual": len(manual_rows), "system": len(system_rows)}]
    for fieldname, label in COMPARE_FIELDS[:-1]:
        rows.append(
            {
                "field": fieldname,
                "label": label,
                "manual": _sum_field(manual_rows, fieldname),
                "system": _sum_field(system_rows, fieldname),
            }
        )
    manual_quantity = _sum_field(manual_rows, "quantity")
    system_quantity = _sum_field(system_rows, "quantity")
    rows.append(
        {
            "field": "weighted_total_unit_rmb",
            "label": "加权综合单价RMB",
            "manual": _safe_div(_sum_field(manual_rows, "total_cost_rmb"), manual_quantity),
            "system": _safe_div(_sum_field(system_rows, "total_cost_rmb"), system_quantity),
        }
    )
    for row in rows:
        row["diff"] = _to_float(row["system"]) - _to_float(row["manual"])
        row["diff_pct"] = _safe_div(row["diff"], _to_float(row["manual"]))
        row["status"] = _diff_status(row["diff"], row["diff_pct"])
    return rows


def write_comparison_xlsx(
    output_path: Path,
    *,
    manual: dict,
    system_result: dict,
    summary_rows: list[dict],
    comparison_rows: list[dict],
    unmatched_system_rows: list[dict],
) -> None:
    if Workbook is None:
        raise RuntimeError("导出对照表需要安装 openpyxl。")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "汇总差异"
    _write_summary_sheet(summary_sheet, manual, system_result, summary_rows)
    _write_detail_sheet(workbook.create_sheet("逐行对照"), comparison_rows)
    _write_unmatched_sheet(workbook.create_sheet("系统未匹配明细"), unmatched_system_rows)
    workbook.save(output_path)


def _write_summary_sheet(sheet, manual: dict, system_result: dict, rows: list[dict]) -> None:
    sheet.append(["项目", "内容"])
    meta_rows = [
        ("人工表来源", f"{manual.get('source_file')} / {manual.get('source_range')}"),
        ("系统批次", system_result.get("business_batch_no") or system_result.get("batch_name")),
        ("Frappe内部记录", system_result.get("batch_doc_name") or system_result.get("batch_name")),
        ("系统版本", system_result.get("version_name")),
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("说明", "本表用于演示对照：人工 Excel 作为历史核算基准，系统结果作为当前试算结果。差异不自动写回。"),
    ]
    for row in meta_rows:
        sheet.append(list(row))
    sheet.append([])
    headers = ["字段", "人工 Excel", "系统试算", "差异(系统-人工)", "差异比例", "状态"]
    sheet.append(headers)
    for row in rows:
        sheet.append([row["label"], row["manual"], row["system"], row["diff"], row["diff_pct"], row["status"]])
    _style_sheet(sheet, header_rows={1, 9}, freeze_cell="A10")
    for row in sheet.iter_rows(min_row=10, min_col=5, max_col=5):
        for cell in row:
            cell.number_format = "0.00%"


def _write_detail_sheet(sheet, rows: list[dict]) -> None:
    headers = [
        "匹配状态",
        "物料编码",
        "人工品名",
        "系统品名",
        "系统规格",
        "人工数量",
        "系统数量",
        "人工货值",
        "系统货值",
        "人工毛重KG",
        "系统毛重KG",
        "人工运费分摊RMB",
        "系统运费分摊RMB",
        "人工清关RMB",
        "系统清关RMB",
        "人工墨西哥内陆/杂费RMB",
        "系统墨西哥内陆/杂费RMB",
        "人工综合成本RMB",
        "系统综合成本RMB",
        "综合成本差异",
        "人工综合单价RMB",
        "系统综合单价RMB",
        "单价差异",
        "人工Excel行",
        "系统行",
        "说明",
    ]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(key) for key in DETAIL_EXPORT_KEYS])
    _style_sheet(sheet, header_rows={1}, freeze_cell="A2")


DETAIL_EXPORT_KEYS = [
    "match_status",
    "material_code",
    "manual_product_name",
    "system_product_name",
    "system_spec_model",
    "manual_quantity",
    "system_quantity",
    "manual_goods_value",
    "system_goods_value",
    "manual_gross_weight_kg",
    "system_gross_weight_kg",
    "manual_freight_alloc_rmb",
    "system_freight_alloc_rmb",
    "manual_mexico_customs_rmb",
    "system_mexico_customs_rmb",
    "manual_mexico_inland_misc_rmb",
    "system_mexico_inland_misc_rmb",
    "manual_total_cost_rmb",
    "system_total_cost_rmb",
    "total_cost_diff",
    "manual_total_unit_rmb",
    "system_total_unit_rmb",
    "total_unit_diff",
    "manual_excel_row",
    "system_row_no",
    "remark",
]


def _write_unmatched_sheet(sheet, rows: list[dict]) -> None:
    headers = ["系统行", "物料编码", "产品名称", "规格型号", "数量", "货值", "毛重KG", "综合成本RMB", "综合单价RMB", "说明"]
    sheet.append(headers)
    for row in rows:
        sheet.append(
            [
                row.get("row_no"),
                row.get("material_code"),
                row.get("product_name"),
                row.get("spec_model"),
                row.get("quantity"),
                row.get("goods_value"),
                row.get("gross_weight_kg"),
                row.get("total_cost_rmb"),
                row.get("total_unit_rmb"),
                "系统有该行，但人工 Excel 基准段未按物料编码/品名匹配到。",
            ]
        )
    _style_sheet(sheet, header_rows={1}, freeze_cell="A2")


def _style_sheet(sheet, *, header_rows: set[int], freeze_cell: str) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(name="Microsoft YaHei", bold=True, color="1F2937")
    normal_font = Font(name="Microsoft YaHei", color="111827")
    border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = header_font if cell.row in header_rows else normal_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            if cell.row in header_rows:
                cell.fill = header_fill
            if isinstance(cell.value, float):
                cell.number_format = "#,##0.0000"
    sheet.freeze_panes = freeze_cell
    sheet.auto_filter.ref = sheet.dimensions
    for index, column_cells in enumerate(sheet.columns, start=1):
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 3, 34)
        sheet.column_dimensions[get_column_letter(index)].width = max(width, 12)


def _build_comparison_row(manual: dict, system: dict | None) -> dict:
    matched = bool(system)
    total_cost_diff = _to_float((system or {}).get("total_cost_rmb")) - _to_float(manual.get("total_cost_rmb"))
    total_unit_diff = _to_float((system or {}).get("total_unit_rmb")) - _to_float(manual.get("total_unit_rmb"))
    return {
        "match_status": "已匹配" if matched else "人工表有，系统未匹配",
        "material_code": manual.get("material_code"),
        "manual_product_name": manual.get("product_name"),
        "system_product_name": (system or {}).get("product_name"),
        "system_spec_model": (system or {}).get("spec_model"),
        "manual_quantity": manual.get("quantity"),
        "system_quantity": (system or {}).get("quantity"),
        "manual_goods_value": manual.get("goods_value"),
        "system_goods_value": (system or {}).get("goods_value"),
        "manual_gross_weight_kg": manual.get("gross_weight_kg"),
        "system_gross_weight_kg": (system or {}).get("gross_weight_kg"),
        "manual_freight_alloc_rmb": manual.get("freight_alloc_rmb"),
        "system_freight_alloc_rmb": (system or {}).get("freight_alloc_rmb"),
        "manual_mexico_customs_rmb": manual.get("mexico_customs_rmb"),
        "system_mexico_customs_rmb": (system or {}).get("mexico_customs_rmb"),
        "manual_mexico_inland_misc_rmb": manual.get("mexico_inland_misc_rmb"),
        "system_mexico_inland_misc_rmb": (system or {}).get("mexico_inland_misc_rmb"),
        "manual_total_cost_rmb": manual.get("total_cost_rmb"),
        "system_total_cost_rmb": (system or {}).get("total_cost_rmb"),
        "total_cost_diff": total_cost_diff if matched else None,
        "manual_total_unit_rmb": manual.get("total_unit_rmb"),
        "system_total_unit_rmb": (system or {}).get("total_unit_rmb"),
        "total_unit_diff": total_unit_diff if matched else None,
        "manual_excel_row": manual.get("manual_excel_row"),
        "system_row_no": (system or {}).get("row_no"),
        "remark": _row_remark(manual, system),
    }


def _row_remark(manual: dict, system: dict | None) -> str:
    if not system:
        return "系统当前批次未匹配到该人工表物料，需确认是否拆分/合并或物料名称不一致。"
    notes = []
    match_reason = system.get("_match_reason") or ""
    if match_reason == "material_code_spec":
        notes.append("按物料编码+规格匹配")
    elif _material_key(manual) != _material_key(system):
        notes.append("编码或品名非完全一致")
    if abs(_to_float(system.get("quantity")) - _to_float(manual.get("quantity"))) > 0.0001:
        notes.append("数量不同")
    if abs(_to_float(system.get("goods_value")) - _to_float(manual.get("goods_value"))) > 0.01:
        notes.append("货值不同")
    if abs(_to_float(system.get("mexico_inland_misc_rmb")) - _to_float(manual.get("mexico_inland_misc_rmb"))) > 0.01:
        notes.append("墨西哥内陆/杂费不同")
    if abs(_to_float(system.get("total_cost_rmb")) - _to_float(manual.get("total_cost_rmb"))) > 0.01:
        notes.append("综合成本不同")
    return "；".join(notes) or "核心字段基本一致"


def _pop_best_system_match(manual: dict, system_pool: list[dict]) -> dict | None:
    exact_key = _material_key(manual)
    for index, row in enumerate(system_pool):
        if _material_key(row) == exact_key:
            matched = system_pool.pop(index)
            matched["_match_reason"] = "exact"
            return matched

    manual_code = _norm(manual.get("material_code"))
    if manual_code:
        matches = [(index, row) for index, row in enumerate(system_pool) if _norm(row.get("material_code")) == manual_code]
        if len(matches) == 1:
            index, _row = matches[0]
            matched = system_pool.pop(index)
            matched["_match_reason"] = "material_code"
            return matched
        if matches:
            manual_name = _norm(manual.get("product_name"))
            for index, row in matches:
                spec_model = _norm(row.get("spec_model"))
                if spec_model and manual_name and (spec_model in manual_name or manual_name in spec_model):
                    matched = system_pool.pop(index)
                    matched["_match_reason"] = "material_code_spec"
                    return matched
            for index, row in matches:
                if manual_name and (manual_name in _norm(row.get("product_name")) or _norm(row.get("product_name")) in manual_name):
                    matched = system_pool.pop(index)
                    matched["_match_reason"] = "material_code_product_name"
                    return matched
    return None


def _read_manual_block(worksheet, start_row: int, customs_no: str, waybill_no: str) -> list[dict]:
    rows = []
    current_customs = customs_no
    current_waybill = waybill_no
    for row_no in range(start_row, worksheet.max_row + 1):
        row_customs = _cell(worksheet, row_no, "I")
        row_waybill = _cell(worksheet, row_no, "J")
        if row_no > start_row and (row_customs or row_waybill):
            if not _matches_token(row_customs, customs_no) and not _matches_token(row_waybill, waybill_no):
                break
        if not _has_data_row(worksheet, row_no):
            break

        row = {"manual_excel_row": row_no}
        for fieldname, column, _label in MANUAL_COLUMNS:
            if column:
                row[fieldname] = _cell(worksheet, row_no, column)
        if row.get("customs_no"):
            current_customs = row["customs_no"]
        else:
            row["customs_no"] = current_customs
        if row.get("waybill_no"):
            current_waybill = row["waybill_no"]
        else:
            row["waybill_no"] = current_waybill
        rows.append(row)
    return rows


def _has_data_row(worksheet, row_no: int) -> bool:
    return any(_cell(worksheet, row_no, column) not in (None, "") for column in ("A", "B", "C", "D", "E", "F", "G", "H"))


def _find_sheet_with_token(workbook, customs_no: str, waybill_no: str):
    for worksheet in workbook.worksheets:
        if _find_start_row(worksheet, customs_no, waybill_no):
            return worksheet
    raise ValueError(f"工作簿中未找到包含 {customs_no} / {waybill_no} 的工作表。")


def _find_start_row(worksheet, customs_no: str, waybill_no: str) -> int:
    for row_no in range(1, worksheet.max_row + 1):
        for col_no in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row_no, col_no).value
            if _matches_token(value, customs_no) or _matches_token(value, waybill_no):
                return row_no
    return 0


def _matches_token(value, token: str) -> bool:
    if value in (None, "") or not token:
        return False
    return _norm(value).find(_norm(token)) >= 0


def _cell(worksheet, row_no: int, column: str):
    return _normalize_value(worksheet[f"{column}{row_no}"].value)


def _normalize_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        text = value.replace("\xa0", " ").strip()
        return text or None
    return value


def _resolve_manual_workbook(file_path: str | None = None) -> Path:
    if file_path:
        path = Path(file_path).expanduser()
        if path.exists():
            return path
        raise FileNotFoundError(f"未找到人工核算 Excel：{file_path}")
    for path in DEFAULT_CANDIDATE_PATHS:
        if path.exists():
            return path
    raise FileNotFoundError("未找到墨西哥进口物料综合成本核算.xlsx，请通过 OVERSEAS_COST_MANUAL_EXCEL 指定。")


def _default_output_path(waybill_no: str) -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"{waybill_no}_系统与人工核算对照_{stamp}.xlsx"
    for parent in (Path("/mnt/c/Users/lin/OneDrive/Desktop"), Path("C:/Users/lin/OneDrive/Desktop"), Path.cwd() / "data"):
        if parent.exists():
            return parent / file_name
    return Path.cwd() / file_name


def _sum_field(rows: list[dict], fieldname: str) -> float:
    return sum(_to_float(row.get(fieldname)) for row in rows)


def _material_key(row: dict) -> tuple[str, str]:
    return (_norm(row.get("material_code")), _norm(row.get("product_name")))


def _norm(value) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"\s+", "", text)


def _safe_div(numerator: float, denominator: float) -> float:
    return numerator / denominator if denominator else 0.0


def _to_float(value, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _diff_status(diff: float, diff_pct: float) -> str:
    if abs(diff) <= 0.01:
        return "一致"
    if abs(diff_pct) <= 0.01:
        return "小差异"
    return "需解释"


def main() -> None:
    result = build_hpcu_manual_comparison_from_env()
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
